from pathlib import Path
from collections import Counter
import sys,subprocess,shutil,tarfile,signal,os,psutil,traceback
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from .errors import CompileError, GitCloneError, RunError, SummaryError


class Ltp_stress():
    def __init__(self, **kwargs):
        self.rpms = {'automake','pkgconf','autoconf','bison','flex','m4','kernel-headers','glibc-headers','findutils','libtirpc','libtirpc-devel','pkg-config','sysstat'}
        self.path = Path('/root/osmts_tmp/ltp_stress')
        self.output_dir = Path('/opt/ltp_stress/output')
        self.directory: Path = kwargs.get('saved_directory') / 'ltp_stress'
        self.compiler: str = kwargs.get('compiler')
        self.remove_osmts_tmp_dir = kwargs.get('remove_osmts_tmp_dir')
        self.result_statuses = ('PASS', 'FAIL', 'CONF', 'BROK', 'WARN', 'TPASS', 'TFAIL', 'TCONF', 'TBROK', 'TWARN')


    # 将进程退出码转为人类可读的描述，负值表示被信号终止
    def _describe_exit_code(self, return_code: int):
        if return_code < 0:
            signal_number = -return_code
            try:
                signal_name = signal.Signals(signal_number).name
            except ValueError:
                signal_name = f'SIGNAL_{signal_number}'
            return f'进程被信号{signal_number}({signal_name})终止'
        return f'进程返回非0退出码:{return_code}'


    # 去除状态 token 首尾标点并转大写，用于与 result_statuses 匹配
    def _normalize_status_token(self, token: str):
        return token.strip().strip(':,[]()').upper()


    # 从行首分离时间戳字段（含可选的 AM/PM），返回时间字符串和剩余字段
    def _split_time_prefix(self, parts: list[str]):
        if not parts or ':' not in parts[0]:
            return '', parts

        time_parts = [parts[0]]
        remaining = parts[1:]
        if remaining and remaining[0].upper() in {'AM', 'PM'}:
            time_parts.append(remaining[0])
            remaining = remaining[1:]
        return ' '.join(time_parts), remaining


    # 判断一行内容是否只是时间戳标记行（非数据行，应跳过）
    def _looks_like_timestamp_marker(self, parts: list[str]):
        if not parts or any(part.lower() == 'device' for part in parts):
            return False
        if len(parts) > 3:
            return False
        return any(':' in part for part in parts)


    # 解析 ltpstress.log，将每条测试记录转为结构化字典并统计各状态数量
    def _parse_ltpstress_log_entries(self, log_path: Path):
        entries = []
        counts = Counter()

        with open(log_path, 'r', encoding='utf-8', errors='replace') as ltpstress_log:
            for line_no, raw_line in enumerate(ltpstress_log, start=1):
                stripped = raw_line.strip()
                if not stripped:
                    continue

                parts = stripped.split()
                status_index = next(
                    (
                        index for index, token in enumerate(parts)
                        if self._normalize_status_token(token) in self.result_statuses
                    ),
                    None,
                )
                if status_index is None:
                    continue

                status = self._normalize_status_token(parts[status_index])
                testcase = parts[0]
                iteration = parts[1] if len(parts) > 1 and parts[1].isdigit() else ''
                exit_value = ''
                note_start_index = status_index + 1

                for index in range(status_index + 1, len(parts)):
                    candidate = parts[index].rstrip(':')
                    if candidate.lstrip('-').isdigit():
                        exit_value = candidate
                        note_start_index = index + 1
                        break

                note = ' '.join(parts[note_start_index:]) if note_start_index < len(parts) else ''
                entries.append({
                    'line_no': line_no,
                    'testcase': testcase,
                    'iteration': iteration,
                    'result': status,
                    'exit_value': exit_value,
                    'note': note,
                    'raw_line': stripped,
                })
                counts[status] += 1

        if not entries:
            raise ValueError('ltpstress.log中未找到可解析的测试结果行')
        return entries, counts


    # 解析 iostat 格式的 iodata 文件，返回表头列表和数据行列表
    def _parse_iodata_rows(self, iodata_path: Path):
        headers = []
        rows = []
        current_time = ''

        with open(iodata_path, 'r', encoding='utf-8', errors='replace') as ltpstress_iodata:
            for raw_line in ltpstress_iodata:
                stripped = raw_line.strip()
                if not stripped or stripped.startswith('Linux '):
                    continue

                parts = stripped.split()
                time_text, remaining = self._split_time_prefix(parts)
                normalized_remaining = [item.lower() for item in remaining]

                if remaining and normalized_remaining[0] == 'device':
                    headers = remaining
                    if time_text:
                        current_time = time_text
                    continue

                if self._looks_like_timestamp_marker(parts):
                    current_time = stripped
                    continue

                if not headers:
                    continue

                if time_text and len(remaining) >= len(headers):
                    current_time = time_text
                    row_values = remaining[:len(headers)]
                elif len(parts) >= len(headers):
                    row_values = parts[:len(headers)]
                else:
                    continue

                rows.append({'Time': current_time, **dict(zip(headers, row_values))})

        if not headers or not rows:
            raise ValueError('ltpstress.iodata中未找到可解析的iostat结果')
        return ['Time'] + headers, rows


    # 解析 sar 输出文本，提取指定指标列的数据行（含 Average 汇总行）
    def _parse_sar_output(self, output_text: str, expected_header_token: str, sheet_desc: str):
        headers = []
        rows = []

        for raw_line in output_text.splitlines():
            stripped = raw_line.strip()
            if not stripped or stripped.startswith('Linux '):
                continue

            parts = stripped.split()
            if stripped.startswith('Average:') or stripped.startswith('平均时间:'):
                if not headers:
                    continue
                row_values = parts[1:]
                if len(row_values) < len(headers):
                    continue
                rows.append({'Time': 'Average', **dict(zip(headers, row_values[:len(headers)]))})
                continue

            time_text, remaining = self._split_time_prefix(parts)
            if not remaining:
                continue

            if expected_header_token in remaining:
                headers = remaining
                continue

            if not headers:
                continue

            if len(remaining) < len(headers):
                continue

            row_values = remaining[:len(headers)]
            rows.append({'Time': time_text, **dict(zip(headers, row_values))})

        if not headers or not rows:
            raise ValueError(f'{sheet_desc}中未找到可解析的sar结果')
        return ['Time'] + headers, rows


    # 按内容自适应调整工作表各列宽度，超长内容启用自动换行
    def _autosize_worksheet(self, worksheet, min_width: int = 12, max_width: int = 60):
        for column_cells in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)
            for cell in column_cells:
                value = '' if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))
                if len(value) > 40:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
            worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, min_width), max_width)


    # 对工作簿中所有工作表执行列宽自适应
    def _autosize_workbook(self, workbook):
        for worksheet in workbook.worksheets:
            self._autosize_worksheet(worksheet)


    def pre_test(self):
        if self.directory.exists():
            shutil.rmtree(self.directory)
        self.directory.mkdir(exist_ok=True, parents=True)
        if self.path.exists():
            shutil.rmtree(self.path)

        # 拉取源码
        repo_url = "https://gitee.com/April_Zhao/ltp_stress.git"
        git_clone = subprocess.run(
            f"cd /root/osmts_tmp/ && git clone {repo_url}",
            shell=True,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.PIPE,
        )
        if git_clone.returncode != 0:
            raise GitCloneError(git_clone.returncode, repo_url, git_clone.stderr.decode('utf-8', errors='replace'))

        # 编译
        if self.compiler == "gcc":
            build = subprocess.run(
                "cd /root/osmts_tmp/ltp_stress && make autotools && ./configure --prefix=/opt/ltp_stress && make -j $(nproc) && make install",
                shell=True,
                stdout=subprocess.DEVNULL,
                stderr=subprocess.PIPE,
            )
        elif self.compiler == "clang":
            build = subprocess.run(
                'cd /root/osmts_tmp/ltp_stress && export CFLAGS="-Wno-error=implicit-function-declaration" && make autotools && CC=clang ./configure --prefix=/opt/ltp_stress && make -j $(nproc) && make install',
                shell=True,
                stdout=subprocess.DEVNULL,
                stderr=subprocess.PIPE,
            )
        else:
            raise ValueError(f"不支持的编译器: {self.compiler}")

        if build.returncode != 0:
            raise CompileError(build.returncode, self.compiler, build.stderr.decode('utf-8', errors='replace'))



    def run_test(self):
        print('ltp_stress测试需要7x24小时,期间请勿中断osmts.')
        if self.output_dir.exists():
            shutil.rmtree(self.output_dir)
        pending_run_error_code = None
        ltpstress_sh = subprocess.Popen(
            "cd /opt/ltp_stress/testscripts && mkdir -p /opt/ltp_stress/output && "
            "./ltpstress.sh -i 3600 -d /opt/ltp_stress/output/ltpstress.data "
            "-I /opt/ltp_stress/output/ltpstress.iodata -l /opt/ltp_stress/output/ltpstress.log "
            "-n -p -S -m 512 -t 168",
            shell=True,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            start_new_session=True
        )

        # 定义Ctrl+C信号处理函数
        def signal_handler(sig, frame):
            print('  osmts检测到Ctrl+C键盘中断信号,正在终止ltp_stress压力测试...')
            try:
                # 尝试终止进程组内所有进程
                os.killpg(os.getpgid(ltpstress_sh.pid), signal.SIGTERM)
            except Exception as e:
                print(f'  终止进程组失败,报错信息{e}',file=sys.stderr)
                print('  尝试递归kill子进程...')
                try:
                    parent = psutil.Process(ltpstress_sh.pid)
                    for child in parent.children(recursive=True):
                        child.kill()
                    parent.kill()
                except psutil.NoSuchProcess:
                    print('  未找到子进程')
            print('  osmts创建的所有子进程均已终止\n当前完整堆栈信息:')
            traceback.print_stack(frame)
            sys.exit(1)

        signal.signal(signal.SIGINT, signal_handler)
        return_code = ltpstress_sh.wait()
        if return_code != 0:
            pending_run_error_code = return_code

        # /opt/ltp/output/里有运行结果
        # ltpstress.log：记录相关日志信息，主要是测试是否通过(pass or fail)
        # ltpstress.data：sar工具记录的日志文件，包括cpu,memory,i/o等
        artifact_paths = {
            'ltpstress.data': self.output_dir / 'ltpstress.data',
            'ltpstress.iodata': self.output_dir / 'ltpstress.iodata',
            'ltpstress.log': self.output_dir / 'ltpstress.log',
        }
        missing_artifacts = [name for name, path in artifact_paths.items() if not path.exists()]
        if missing_artifacts:
            raise RunError(0, f"ltp_stress测试结束后缺少结果文件:{', '.join(missing_artifacts)}")

        for artifact_name, artifact_path in artifact_paths.items():
            shutil.copy2(artifact_path, self.directory / artifact_name)

        with tarfile.open(self.directory / 'ltpstress.tar.xz','w:xz') as tar:
            for artifact_name in artifact_paths:
                tar.add(self.directory / artifact_name, arcname=artifact_name)
        # ---------------------------------------------------------------

        # 分析ltpstress.log文件,进行统计
        wb = Workbook()
        ws = wb.active
        ws.title = 'ltp stress report'
        ws.append(['Line', 'Testcase', 'Iteration', 'Result', 'Exit Value', 'Note', 'Raw Line'])
        log_entries, result_counts = self._parse_ltpstress_log_entries(self.directory / 'ltpstress.log')
        for entry in log_entries:
            ws.append([
                entry['line_no'],
                entry['testcase'],
                entry['iteration'],
                entry['result'],
                entry['exit_value'],
                entry['note'],
                entry['raw_line'],
            ])

        wb.create_sheet(title='ltp stress summary')
        ws_summary = wb['ltp stress summary']
        ws_summary.append(['Metric', 'Value', 'Note'])
        ws_summary.append(['total parsed result lines', len(log_entries), 'ltpstress.log中成功解析的结果行数'])
        for status in self.result_statuses:
            if result_counts.get(status):
                ws_summary.append([f'{status} count', result_counts[status], ''])
        fail_like_entries = [entry for entry in log_entries if 'FAIL' in entry['result'] or 'BROK' in entry['result']]
        ws_summary.append(['fail/brok lines', len(fail_like_entries), 'FAIL、TFAIL、BROK、TBROK总行数'])
        ws_summary.append([
            'unique fail/brok testcases',
            len({entry['testcase'] for entry in fail_like_entries}),
            '去重后的失败/阻塞测试项数量',
        ])


        # 分析ltpstress.iodata
        wb.create_sheet(title='ltp stress iodata')
        ws = wb['ltp stress iodata']
        iodata_headers, iodata_rows = self._parse_iodata_rows(self.directory / 'ltpstress.iodata')
        ws.append(iodata_headers)
        for row in iodata_rows:
            ws.append([row.get(header, '') for header in iodata_headers])

        # 分析sar -u结果 (CPU使用率历史数据)
        sar_u = subprocess.run(
            f"sar -u -f {self.directory / 'ltpstress.data'}",
            shell=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.DEVNULL,
        )
        cpu_output = sar_u.stdout.decode('utf-8', errors='replace')
        with open(self.directory / 'sar_cpu.txt', 'w', encoding='utf-8') as cpu_log:
            cpu_log.write(cpu_output)
        wb.create_sheet(title='sar cpu')
        ws = wb['sar cpu']
        cpu_headers, cpu_rows = self._parse_sar_output(cpu_output, '%user', 'sar cpu输出')
        ws.append(cpu_headers)
        for row in cpu_rows:
            ws.append([row.get(header, '') for header in cpu_headers])

        # 分析sar -r结果 (Memory使用率历史数据)
        sar_r = subprocess.run(
            f"sar -r -f {self.directory / 'ltpstress.data'}",
            shell=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.DEVNULL,
        )
        mem_output = sar_r.stdout.decode('utf-8', errors='replace')
        with open(self.directory / 'sar_memory.txt', 'w', encoding='utf-8') as mem_log:
            mem_log.write(mem_output)
        wb.create_sheet(title='sar memory')
        ws = wb['sar memory']
        mem_headers, mem_rows = self._parse_sar_output(mem_output, 'kbmemfree', 'sar memory输出')
        ws.append(mem_headers)
        for row in mem_rows:
            ws.append([row.get(header, '') for header in mem_headers])

        self._autosize_workbook(wb)
        wb.save(self.directory / 'ltp_stress_report.xlsx')
        if pending_run_error_code is not None:
            result_log = self.output_dir / 'ltpstress.log'
            error_message = (
                f"ltp_stress压力测试{self._describe_exit_code(pending_run_error_code)}."
                f"原始日志请查看:{result_log}. "
                f"已保存的测试产物目录:{self.directory}"
            )
            raise RunError(pending_run_error_code, error_message)


    def post_test(self):
        # ltp_stress运行后/tmp/目录下会产生缓存文件占满整个tmpfs,因此务必清除
        subprocess.run(
            "rm -rf /tmp/ltpstress*",
            shell=True,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
        )



    def run(self):
        print("开始进行ltp_stress测试")
        need_cleanup = False
        try:
            self.pre_test()
            need_cleanup = True
            try:
                self.run_test()
            except RunError:
                raise
            except Exception as e:
                summary_error_log = self.directory / 'ltp_stress_summary_error.log'
                with open(summary_error_log, 'w', encoding='utf-8') as log:
                    log.write(str(e))
                raise SummaryError(summary_error_log)
        finally:
            if need_cleanup:
                self.post_test()
        print("ltp_stress测试结束")

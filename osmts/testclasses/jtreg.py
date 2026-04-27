import re
import shutil
import subprocess
import zipfile
import os
import tarfile
import shlex
import time
import requests
from pathlib import Path
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Alignment

from .errors import DnfError, DefaultError


"""
使用 Jtreg 对 OpenJDK 执行测试回归测试
根据系统安装的 OpenJDK 版本自动下载对应源码包和 jtreg 工具
"""

headers = {
    'Connection': 'keep-alive',
    'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64; rv:136.0) Gecko/20100101 Firefox/136.0',
    'Referer': 'https://gitee.com/April_Zhao/osmts'
}

JDK_RPM_PATTERNS = {
    8: 'java-1.8.0-openjdk*',
    11: 'java-11-openjdk*',
    17: 'java-17-openjdk*',
}

JDK_GITHUB_REPOS = {
    8: 'openjdk/jdk8u-dev',
    11: 'openjdk/jdk11u-dev',
    17: 'openjdk/jdk17u-dev',
}

JTREG_SHIPILEV_URL = {
    8: 'https://builds.shipilev.net/jtreg/jtreg4.2-b16.zip',
    11: 'https://builds.shipilev.net/jtreg/jtreg-7.3.1+1.zip',
    17: 'https://builds.shipilev.net/jtreg/jtreg-7.3.1+1.zip',
}

JTREG_EXPECTED_DIR = {
    8: 'jtreg-4.2',
    11: 'jtreg-7.3.1',
    17: 'jtreg-7.3.1',
}


def install_rpm(package_name):
    try:
        subprocess.run(
            f"dnf install -y {package_name}",
            shell=True, check=True,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.PIPE,
        )
    except subprocess.CalledProcessError as e:
        stderr_msg = e.stderr.decode('utf-8') if e.stderr else ''
        raise DnfError(e.returncode, stderr_msg)


def remove_rpm(package_name):
    try:
        subprocess.run(
            f"dnf remove -y {package_name}",
            shell=True, check=True,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.PIPE,
        )
    except subprocess.CalledProcessError as e:
        stderr_msg = e.stderr.decode('utf-8') if e.stderr else ''
        raise DnfError(e.returncode, stderr_msg)


def clean_java_environment():
    for pattern in JDK_RPM_PATTERNS.values():
        remove_rpm(pattern)


# 根据版本号分量构造对应的 OpenJDK 源码 tag 字符串
def _build_jdk_tag(major, ver1, ver2, ver3, ver4, ver5):
    if major == 8:
        return f"jdk8u{int(ver4):03d}-b{ver5}"
    else:
        patch = ver4 if ver4 else ver3
        return f"jdk-{ver1}.{ver2}.{patch}+{ver5}"


def detect_java_version(major_ver):
    result = subprocess.run(
        "java -version",
        shell=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
    )
    stdout = result.stdout.decode('utf-8') if result.stdout else ''
    stderr = result.stderr.decode('utf-8') if result.stderr else ''
    output = stdout + stderr
    print(f"[jtreg] java -version 原始输出:\n{output.strip() if output.strip() else '(空)'}")

    line1 = output.splitlines()[0] if output else ''

    ver_match = re.search(r'"(\d+)\.(\d+)\.(\d+)(?:_(\d+))?"', line1)
    if not ver_match:
        print(f"[jtreg] 未从 java -version 第一行解析出版本号，跳过 JDK {major_ver}")
        return None

    ver1, ver2, ver3, ver4 = ver_match.groups()
    if ver1 == '1' and ver2 == '8':
        detected_major = 8
    else:
        detected_major = int(ver1)

    has_underscore = '_' in line1

    b_match = re.search(r'\(build\s+[\d.]+_(\d+)-b(\d+)\)', output)
    if b_match:
        b4, b5 = b_match.groups()
        ver5 = b5
        if ver4 is None:
            ver4 = b4
    else:
        b_match2 = re.search(r'\(build\s+([\d.]+)\+(\d+)\)', output)
        if b_match2:
            b_ver, b5 = b_match2.groups()
            ver5 = b5
            if ver4 is None:
                b_parts = b_ver.split('.')
                if detected_major == 8 and len(b_parts) >= 3:
                    ver4 = b_parts[2]
                else:
                    ver4 = '0'
        else:
            ver5 = '0'
            ver4 = ver4 or '0'

    update = ver4 if ver4 else '0'
    full_version = (f"{ver1}.{ver2}.{ver3}_{update}-b{ver5}"
                    if update != '0'
                    else f"{ver1}.{ver2}.{ver3}-b{ver5}")

    print(f"[jtreg] 解析到版本: {full_version}, 映射为主版本: {detected_major}")

    if detected_major != major_ver:
        print(f"[jtreg] 检测到主版本 {detected_major}，与预期 {major_ver} 不符，跳过")
        return None

    tag_patch = ver3 if ver4 else ver3
    tag_update = update if has_underscore else ver3
    tag = _build_jdk_tag(detected_major, ver1, ver2, tag_patch, tag_update, ver5)
    repo = JDK_GITHUB_REPOS[detected_major]
    src_url = f"https://github.com/{repo}/archive/refs/tags/{tag}.zip"
    print(f"[jtreg] 匹配到 OpenJDK {detected_major} (build: {full_version}), tag: {tag}")
    print(f"[jtreg]   源码下载地址: {src_url}")
    return {
        'full': full_version,
        'tag': tag,
        'src_url': src_url,
        'src_repo': repo,
        'rpm': JDK_RPM_PATTERNS[detected_major],
        'jt_home': None,
        'test_dir': None,
    }


# 在指定目录下查找以 jtreg 开头的子目录
def _find_jtreg_dir(base_path):
    for item in base_path.iterdir():
        if item.is_dir() and item.name.startswith('jtreg'):
            return item
    return None


# 带重试机制地下载 zip 文件并解压到指定目录
def _download_and_extract_zip(url, extract_to, progress_label, timeout, max_retries=3):
    archive_path = extract_to / f".{progress_label.replace(' ', '_').replace('/', '_')}.zip"
    last_error = None

    for attempt in range(1, max_retries + 1):
        if archive_path.exists():
            archive_path.unlink()

        print(f"[jtreg] {progress_label}: {url}")
        if attempt > 1:
            print(f"[jtreg] 第 {attempt}/{max_retries} 次重试下载 {progress_label}")

        try:
            with requests.get(url, headers=headers, timeout=timeout, stream=True) as response:
                response.raise_for_status()
                total = int(response.headers.get('content-length', 0))
                downloaded = 0

                with open(archive_path, 'wb') as archive_file:
                    for chunk in response.iter_content(chunk_size=1024 * 1024):
                        if not chunk:
                            continue
                        archive_file.write(chunk)
                        downloaded += len(chunk)
                        if total > 0:
                            pct = downloaded * 100 // total
                            print(f"\r[jtreg] {progress_label} 下载进度: {pct}%", end='', flush=True)

                if total > 0:
                    print()

            if not zipfile.is_zipfile(archive_path):
                raise zipfile.BadZipFile(f"{archive_path} 不是有效的 zip 文件")

            archive_size_kb = archive_path.stat().st_size // 1024
            print(f"[jtreg] {progress_label} 下载完成 ({archive_size_kb} KB), 开始解压...")
            with zipfile.ZipFile(archive_path, 'r') as zf:
                zf.extractall(path=extract_to)
            print(f"[jtreg] {progress_label} 解压完成")
            archive_path.unlink(missing_ok=True)
            return
        except (requests.RequestException, zipfile.BadZipFile, OSError) as e:
            last_error = e
            print(f"[jtreg] {progress_label} 下载失败: {e}")
            archive_path.unlink(missing_ok=True)
            if attempt < max_retries:
                wait_seconds = min(5 * attempt, 15)
                print(f"[jtreg] {wait_seconds} 秒后重试...")
                time.sleep(wait_seconds)

    raise last_error


# 下载并解压指定版本的 JDK 源码
def _download_jdk_source(url, extract_to, major_ver):
    _download_and_extract_zip(
        url=url,
        extract_to=extract_to,
        progress_label=f"下载 JDK {major_ver} 源码",
        timeout=(30, 600),
    )


# 下载并解压指定版本对应的 jtreg 工具包
def _download_jtreg(url, extract_to, major_ver):
    _download_and_extract_zip(
        url=url,
        extract_to=extract_to,
        progress_label=f"下载 jtreg 工具 (JDK {major_ver})",
        timeout=(30, 300),
    )
    jt_bin = None
    for root, dirs, files in os.walk(extract_to):
        for f in files:
            if f == 'jtreg' and 'bin' in root:
                jt_bin = Path(root) / f
                break
        if jt_bin:
            break
    if jt_bin and jt_bin.exists():
        jt_bin.chmod(0o755)
        print(f"[jtreg] 已设置 {jt_bin} 执行权限")


class Jtreg:
    def __init__(self, **kwargs):
        self.rpms = {'subversion', 'screen', 'samba', 'samba-client', 'gdb',
                     'automake', 'lrzsz', 'expect', 'libX11*', 'libxt*', 'libXtst*',
                     'libXt*', 'libXrender*', 'cache*', 'cups*', 'freetype*',
                     'mercurial', 'numactl', 'vim', 'tar', 'dejavu-fonts',
                     'unix2dos', 'dos2unix', 'bc', 'lsof', 'net-tools'}
        self.path = Path('/root/osmts_tmp/jtreg')
        self.directory: Path = kwargs.get('saved_directory') / 'jtreg'

    # 初始化结果目录和工作目录，清理上次遗留的临时文件
    def _setup_workdir(self):
        if not self.directory.exists():
            self.directory.mkdir(parents=True)
        if self.path.exists():
            shutil.rmtree(self.path)
        self.path.mkdir(parents=True)

    # 在工作目录中定位 jtreg 可执行目录和 JDK 源码目录，填充 info 字典
    def _resolve_paths(self, info, major_ver):
        print(f"[jtreg] 解压后 {self.path} 目录内容: {list(self.path.iterdir())}")

        expected_dir = JTREG_EXPECTED_DIR[major_ver]
        jt_home = self.path / expected_dir
        if not jt_home.exists():
            found = _find_jtreg_dir(self.path)
            if found:
                jt_home = found
                print(f"[jtreg] 未找到预期目录 {expected_dir}, "
                      f"已找到 jtreg 目录: {jt_home.name}")
            else:
                raise FileNotFoundError(
                    f"无法在 {self.path} 中找到 jtreg 目录"
                )

        repo_name = info['src_repo'].split('/')[1]
        repo_dir = None
        for item in self.path.iterdir():
            if item.is_dir() and repo_name in item.name:
                repo_dir = item
                break

        if repo_dir is None:
            raise FileNotFoundError(
                f"无法在 {self.path} 中找到源码目录 (关键词: {repo_name})"
            )

        if not (repo_dir / 'test').exists():
            raise FileNotFoundError(
                f"源码目录 {repo_dir} 中不存在 test 子目录"
            )

        info['jt_home'] = str(jt_home)
        info['test_dir'] = str(repo_dir)
        info['repo_dir'] = repo_dir
        print(f"[jtreg] JDK {major_ver} 路径解析完成:")
        print(f"  JT_HOME = {jt_home}")
        print(f"  test_dir (repo root) = {repo_dir}")

    # 返回指定 JDK 主版本对应的 jtreg 测试集规格字符串
    def _get_test_spec(self, major_ver):
        specs = {
            8: 'hotspot/test:hotspot_tier1 langtools/test:langtools_tier1 jdk/test:jdk_tier1',
            11: 'test/langtools:tier1 test/hotspot/jtreg:tier1 test/jdk:tier1 test/jaxp:tier1',
            17: 'test/hotspot/jtreg:tier1 test/langtools:tier1 test/jdk:tier1 test/jaxp:tier1 test/lib-test:tier1',
        }
        return specs.get(major_ver, '')

    # 创建并返回指定 JDK 版本的 JTwork 和 JTreport 输出目录路径
    def _prepare_version_output_dirs(self, major_ver):
        version_root = self.path / f'jdk{major_ver}'
        if version_root.exists():
            shutil.rmtree(version_root)
        version_root.mkdir(parents=True, exist_ok=True)
        return {
            'root': version_root,
            'jtwork': version_root / 'JTwork',
            'jtreport': version_root / 'JTreport',
        }

    # 构造并执行 jtreg 命令，将日志输出到文件，返回退出码
    def _run_jtreg(self, version_info, test_spec, major_ver):
        jt_home = Path(version_info['jt_home'])
        test_dir = Path(version_info['test_dir'])
        output_dirs = self._prepare_version_output_dirs(major_ver)
        log_path = self.directory / f'OpenJDK{major_ver}.log'
        cmd = [
            str(jt_home / 'bin' / 'jtreg'),
            '-va',
            '-ignore:quiet',
            '-jit',
            '-conc:auto',
            '-timeout:16',
            '-tl:3590',
            f"-w:{output_dirs['jtwork']}",
            f"-r:{output_dirs['jtreport']}",
            *shlex.split(test_spec),
        ]
        env = os.environ.copy()
        env['JT_HOME'] = str(jt_home)

        print(f"[jtreg] 执行 jtreg 命令:")
        print(f"  JT_HOME = {jt_home}")
        print(f"  test_dir = {test_dir}")
        print(f"  test_spec = {test_spec}")
        print(f"  JTreport = {output_dirs['jtreport']}")
        print(f"  JTwork = {output_dirs['jtwork']}")

        with open(log_path, 'w', encoding='utf-8') as log:
            jtreg = subprocess.run(
                cmd,
                cwd=test_dir,
                env=env,
                stdout=log,
                stderr=subprocess.STDOUT,
                text=True,
            )

        version_info['jt_home'] = jt_home
        version_info['test_dir'] = test_dir
        version_info['jtreport'] = output_dirs['jtreport']
        version_info['jtwork'] = output_dirs['jtwork']
        version_info['log_path'] = log_path
        version_info['returncode'] = jtreg.returncode
        version_info['command'] = ' '.join(shlex.quote(part) for part in cmd)
        return jtreg.returncode

    # 将 jtreg 生成的 summary.txt 复制到结果目录
    def _copy_summary(self, version_info, major_ver):
        summary_source = version_info['jtreport'] / 'text' / 'summary.txt'
        if not summary_source.exists():
            raise DefaultError(
                f"jtreg测试出错.OpenJDK {major_ver} 未生成JTreport/text/summary.txt,"
                f"jtreg返回值:{version_info['returncode']}.详细信息请查看:{version_info['log_path']}"
            )

        summary_target = self.directory / f'OpenJDK{major_ver}_summary.txt'
        shutil.copy2(summary_source, summary_target)
        version_info['summary_path'] = summary_target
        print(f"[jtreg] 已保存 summary: {summary_target}")
        return summary_target

    # 将 JTreport 和 JTwork 目录打包为 tar.gz 归档文件
    def _package_output(self, version_info, major_ver):
        jtreport = version_info['jtreport']
        jtwork = version_info['jtwork']
        pkg_name = self.directory / f'OpenJDK{major_ver}_jtreg_output.tar.gz'

        items_to_pack = []
        if jtreport.exists():
            items_to_pack.append(jtreport)
        if jtwork.exists():
            items_to_pack.append(jtwork)

        if not items_to_pack:
            print(f"[jtreg] 未找到 JTreport 或 JTwork，跳过打包")
            return

        print(f"[jtreg] 正在打包 JTreport/JTwork 到 {pkg_name}")
        with tarfile.open(pkg_name, 'w:gz') as tar:
            for item in items_to_pack:
                tar.add(item, arcname=item.name)
        print(f"[jtreg] 打包完成: {pkg_name}")
        version_info['package_path'] = pkg_name
        return pkg_name

    # 解析 jtreg summary.txt，将每行测试结果转为含状态和详情的字典
    def _parse_summary_file(self, summary_path: Path):
        rows = []
        for raw_line in summary_path.read_text(encoding='utf-8', errors='replace').splitlines():
            line = raw_line.rstrip()
            if not line:
                continue

            match = re.match(r'^(?P<test_name>.+?)\s{2,}(?P<result_text>.+)$', line)
            if not match:
                rows.append({
                    'test_name': line,
                    'status': 'UNPARSED',
                    'detail': '',
                })
                continue

            test_name = match.group('test_name').strip()
            result_text = match.group('result_text').strip()
            result_match = re.match(r'^(?P<status>[^.]+)\.\s*(?P<detail>.*)$', result_text)
            if result_match:
                status = result_match.group('status').strip()
                detail = result_match.group('detail').strip()
            else:
                status = result_text
                detail = ''

            rows.append({
                'test_name': test_name,
                'status': status,
                'detail': detail,
            })

        return rows

    # 将所有 JDK 版本的 jtreg 测试结果汇总写入 Excel 文件
    def _save_excel(self, version_results):
        wb = Workbook()
        ws_overview = wb.active
        ws_overview.title = 'overview'
        ws_overview.append([
            'JDK',
            'build',
            'jtreg returncode',
            'total',
            'passed',
            'failed',
            'error',
            'other',
            'summary file',
            'log file',
            'artifact',
            'note',
        ])

        for result in version_results:
            counter = Counter(row['status'] for row in result['rows'])
            total = len(result['rows'])
            passed = counter.get('Passed', 0)
            failed = counter.get('Failed', 0)
            error = counter.get('Error', 0)
            other = total - passed - failed - error

            ws_overview.append([
                f"OpenJDK{result['major_ver']}",
                result['build'],
                result['returncode'],
                total,
                passed,
                failed,
                error,
                other,
                str(result['summary_path']),
                str(result['log_path']),
                str(result['package_path']),
                'jtreg非0返回值通常表示存在失败/错误用例, 以summary.txt统计为准',
            ])

            ws_detail = wb.create_sheet(title=f"OpenJDK{result['major_ver']}")
            ws_detail.append(['test case', 'status', 'detail'])
            for row in result['rows']:
                ws_detail.append([row['test_name'], row['status'], row['detail']])

            ws_detail.freeze_panes = 'A2'
            ws_detail.column_dimensions['A'].width = 72
            ws_detail.column_dimensions['B'].width = 14
            ws_detail.column_dimensions['C'].width = 52
            for cells in ws_detail.iter_rows():
                for cell in cells:
                    cell.alignment = Alignment(vertical='top', wrap_text=True)

        ws_overview.freeze_panes = 'A2'
        ws_overview.column_dimensions['A'].width = 12
        ws_overview.column_dimensions['B'].width = 24
        ws_overview.column_dimensions['C'].width = 16
        ws_overview.column_dimensions['D'].width = 10
        ws_overview.column_dimensions['E'].width = 10
        ws_overview.column_dimensions['F'].width = 10
        ws_overview.column_dimensions['G'].width = 10
        ws_overview.column_dimensions['H'].width = 10
        ws_overview.column_dimensions['I'].width = 34
        ws_overview.column_dimensions['J'].width = 30
        ws_overview.column_dimensions['K'].width = 34
        ws_overview.column_dimensions['L'].width = 44
        for cells in ws_overview.iter_rows():
            for cell in cells:
                cell.alignment = Alignment(vertical='top', wrap_text=True)

        wb.save(self.directory / 'jtreg.xlsx')

    def run(self):
        print('开始进行jtreg测试')
        self._setup_workdir()
        version_results = []

        for major_ver in sorted(JDK_RPM_PATTERNS.keys()):
            print(f"[jtreg] ========== JDK {major_ver} 测试开始 ==========")

            rpm = JDK_RPM_PATTERNS[major_ver]
            installed = False
            version_error = None
            try:
                print(f"[jtreg] 第1步: 安装 {rpm}")
                install_rpm(rpm)
                installed = True

                print(f"[jtreg] 第2步: 检测 Java 版本")
                info = detect_java_version(major_ver)
                if info is None:
                    print(f"[jtreg] JDK {major_ver} 版本检测失败，跳过")
                    continue

                print(f"[jtreg] 第3步: 下载源码和 jtreg")
                _download_jdk_source(info['src_url'], self.path, major_ver)
                _download_jtreg(JTREG_SHIPILEV_URL[major_ver], self.path, major_ver)

                print(f"[jtreg] 第4步: 解析路径")
                self._resolve_paths(info, major_ver)

                print(f"  开始进行OpenJDK {major_ver} 测试 (build: {info['full']})")
                self._run_jtreg(info, self._get_test_spec(major_ver), major_ver)
                print(f"  OpenJDK {major_ver} 测试结束")

                self._copy_summary(info, major_ver)
                self._package_output(info, major_ver)

                version_results.append({
                    'major_ver': major_ver,
                    'build': info['full'],
                    'returncode': info['returncode'],
                    'summary_path': info['summary_path'],
                    'log_path': info['log_path'],
                    'package_path': info['package_path'],
                    'rows': self._parse_summary_file(info['summary_path']),
                })
                self._save_excel(version_results)
            except Exception as e:
                version_error = e
                raise
            finally:
                if installed:
                    print(f"[jtreg] 第5步: 卸载 {rpm}")
                    try:
                        remove_rpm(rpm)
                    except Exception as cleanup_error:
                        if version_error is None:
                            raise
                        print(f"[jtreg] 卸载 {rpm} 失败: {cleanup_error}")
                print(f"[jtreg] ========== JDK {major_ver} 测试结束 ==========\n")

        print('jtreg测试结束')

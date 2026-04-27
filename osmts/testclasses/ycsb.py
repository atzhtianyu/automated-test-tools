from openpyxl.workbook import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from pystemd.systemd1 import Unit
from pathlib import Path
import re,time,shlex
import subprocess,shutil

from .errors import DefaultError,GitCloneError,RunError,SummaryError



class YCSB: # Yahoo！Cloud Serving Benchmark
    def __init__(self, **kwargs):
        self.rpms = {'redis','java','maven'}
        self.believe_tmp: bool = kwargs.get('believe_tmp')
        self.path = Path('/root/osmts_tmp/ycsb')
        self.directory: Path = kwargs.get('saved_directory') / 'ycsb'
        self.ycsb:Path = self.path / 'bin/ycsb'
        self.workloada:Path = self.path / 'workloads/workloada'
        self.load_result:str = ''
        self.test_result:str = ''
        self.load_command = "bin/ycsb load redis -threads 100 -P workloads/workloada"
        self.run_command = "bin/ycsb run redis -threads 100 -P workloads/workloada"


    # 读取 YCSB workload 配置文件，返回键值对字典
    def _load_workload_config(self):
        config = {}
        if not self.workloada.exists():
            return config

        with open(self.workloada) as workload:
            for line in workload:
                stripped = line.strip()
                if not stripped or stripped.startswith('#') or '=' not in stripped:
                    continue
                key, value = stripped.split('=', 1)
                config[key.strip()] = value.strip()
        return config


    # 将指定键值覆写到 workload 配置文件，保留其余行不变
    def _save_workload_config(self, overrides):
        lines = []
        updated_keys = set()

        if self.workloada.exists():
            with open(self.workloada) as workload:
                original_lines = workload.readlines()
        else:
            original_lines = []

        for raw_line in original_lines:
            stripped = raw_line.strip()
            if not stripped or stripped.startswith('#') or '=' not in raw_line:
                lines.append(raw_line)
                continue

            key = raw_line.split('=', 1)[0].strip()
            if key in overrides:
                if key in updated_keys:
                    continue
                line_ending = '\n' if raw_line.endswith('\n') else ''
                lines.append(f"{key}={overrides[key]}{line_ending}")
                updated_keys.add(key)
            else:
                lines.append(raw_line)

        if lines and not lines[-1].endswith('\n'):
            lines[-1] = lines[-1] + '\n'

        for key, value in overrides.items():
            if key not in updated_keys:
                lines.append(f"{key}={value}\n")

        with open(self.workloada, 'w') as workload:
            workload.writelines(lines)


    # 从 YCSB 命令行字符串中提取 -threads、-P、-target 等参数
    def _parse_command_args(self, command: str):
        args = {}
        tokens = shlex.split(command)
        index = 0
        while index < len(tokens):
            token = tokens[index]
            if token in {'-threads', '-P', '-target'} and index + 1 < len(tokens):
                args[token] = tokens[index + 1]
                index += 2
                continue
            index += 1
        return args


    # 将字符串指标值自动转换为 int、float 或保持原字符串
    def _parse_metric_value(self, value: str):
        value = value.strip()
        if re.fullmatch(r"-?\d+", value):
            return int(value)
        if re.fullmatch(r"-?\d+\.\d+", value):
            return float(value)
        return value


    # 将浮点比率格式化为四位小数的百分比字符串
    def _format_ratio_percent(self, ratio):
        if isinstance(ratio, (int, float)):
            return f"{ratio:.4%}"
        return ''


    # 解析 YCSB 日志中的 [section], metric, value 格式记录，返回结构化列表
    def _parse_records(self):
        records = []
        for line in self.test_result.splitlines():
            match = re.match(r"^\[(?P<section>[^\]]+)\], (?P<metric>[^,]+), (?P<value>.+)$", line.strip())
            if match:
                records.append({
                    'section': match.group('section').strip(),
                    'metric': match.group('metric').strip(),
                    'value': self._parse_metric_value(match.group('value')),
                })

        if not records:
            raise ValueError('ycsb日志中未找到可解析的指标')
        return records


    # 将记录列表按 section 分组，返回 {section: {metric: value}} 嵌套字典
    def _group_records(self, records):
        grouped = {}
        for record in records:
            section = record['section']
            grouped.setdefault(section, {})
            grouped[section][record['metric']] = record['value']
        return grouped


    # 从分组记录中提取 GC 统计信息，构建 Excel 所需的行数据列表
    def _build_gc_rows(self, grouped_records):
        total_gc = {
            'collector': 'TOTAL',
            'count': '',
            'time_ms': '',
            'time_percent': '',
        }
        collectors = {}

        for section, metrics in grouped_records.items():
            if section == 'TOTAL_GCs':
                total_gc['count'] = metrics.get('Count', '')
            elif section == 'TOTAL_GC_TIME':
                total_gc['time_ms'] = metrics.get('Time(ms)', '')
            elif section == 'TOTAL_GC_TIME_%':
                total_gc['time_percent'] = metrics.get('Time(%)', '')
            elif section.startswith('TOTAL_GC_TIME_%_'):
                collector = section.removeprefix('TOTAL_GC_TIME_%_')
                collectors.setdefault(collector, {'collector': collector, 'count': '', 'time_ms': '', 'time_percent': ''})
                collectors[collector]['time_percent'] = metrics.get('Time(%)', '')
            elif section.startswith('TOTAL_GC_TIME_'):
                collector = section.removeprefix('TOTAL_GC_TIME_')
                collectors.setdefault(collector, {'collector': collector, 'count': '', 'time_ms': '', 'time_percent': ''})
                collectors[collector]['time_ms'] = metrics.get('Time(ms)', '')
            elif section.startswith('TOTAL_GCS_'):
                collector = section.removeprefix('TOTAL_GCS_')
                collectors.setdefault(collector, {'collector': collector, 'count': '', 'time_ms': '', 'time_percent': ''})
                collectors[collector]['count'] = metrics.get('Count', '')

        rows = [total_gc]
        rows.extend(collectors[name] for name in sorted(collectors))
        return rows


    # 收集所有操作 section 中出现的 Return= 指标名，Return=OK 排在最前
    def _collect_return_metric_names(self, grouped_records):
        return_metric_names = set()
        for section, metrics in grouped_records.items():
            if section == 'OVERALL' or section.startswith('TOTAL_'):
                continue
            for metric_name in metrics:
                if metric_name.startswith('Return='):
                    return_metric_names.add(metric_name)

        return sorted(
            return_metric_names,
            key=lambda name: (name != 'Return=OK', name),
        )


    # 构建各操作的 Excel 行数据，包含吞吐量、延迟和各 Return 状态计数
    def _build_operation_rows(self, grouped_records, return_metric_names):
        operation_rows = []
        for section, metrics in grouped_records.items():
            if section == 'OVERALL' or section.startswith('TOTAL_'):
                continue

            operations = metrics.get('Operations', '')
            return_metrics = {metric_name: metrics.get(metric_name, '') for metric_name in return_metric_names}
            return_ok = return_metrics.get('Return=OK', '')
            success_rate = ''
            if isinstance(operations, int) and operations > 0 and isinstance(return_ok, int):
                success_rate = round(return_ok / operations, 4)
            non_ok_return_total = sum(
                count for metric_name, count in return_metrics.items()
                if metric_name != 'Return=OK' and isinstance(count, int)
            )

            operation_rows.append({
                'operation': section,
                'operations': operations,
                'return_metrics': return_metrics,
                'return_ok': return_ok,
                'non_ok_return_total': non_ok_return_total,
                'success_rate': success_rate,
                'avg_latency_us': metrics.get('AverageLatency(us)', ''),
                'min_latency_us': metrics.get('MinLatency(us)', ''),
                'max_latency_us': metrics.get('MaxLatency(us)', ''),
                'p50_latency_us': metrics.get('50thPercentileLatency(us)', ''),
                'p95_latency_us': metrics.get('95thPercentileLatency(us)', ''),
                'p99_latency_us': metrics.get('99thPercentileLatency(us)', ''),
            })

        if not operation_rows:
            raise ValueError('ycsb日志中未找到操作类指标')
        return operation_rows


    # 执行 YCSB 命令并保存输出到日志文件
    def _run_ycsb_command(self, command: str, log_filename: str, error_msg: str) -> str:
        try:
            result = subprocess.run(
                command, cwd=self.path, shell=True, check=True,
                stdout=subprocess.PIPE, stderr=subprocess.PIPE,
            )
        except subprocess.CalledProcessError as e:
            raise DefaultError(f"{error_msg},报错信息:{e.stderr.decode('utf-8')}")

        output = result.stdout.decode('utf-8')
        stderr_text = result.stderr.decode('utf-8')
        if stderr_text:
            output += f"\n[stderr]\n{stderr_text}"

        with open(self.directory / log_filename, 'w') as log:
            log.write(output)
        return output


    def pre_test(self):
        self.redis:Unit = Unit(b'redis.service',_autoload=True)
        try:
            self.redis.Unit.Start(b'replace')
        except Exception:
            self.redis.Unit.Start(b'replace')
        time.sleep(5)
        if self.redis.Unit.ActiveState != b'active':
            time.sleep(5)
            if self.redis.Unit.ActiveState != b'active':
                raise DefaultError("redis.service开启失败,退出测试.")

        if self.directory.exists():
            shutil.rmtree(self.directory)
        self.directory.mkdir(parents=True)

        if self.path.exists() and self.believe_tmp:
            pass
        else:
            shutil.rmtree(self.path,ignore_errors=True)
            try:
                subprocess.run(
                    "git clone https://gitee.com/zhtianyu/ycsb.git",
                    cwd="/root/osmts_tmp",
                    shell=True,check=True,
                    stdout=subprocess.DEVNULL,
                    stderr=subprocess.PIPE,
                )
            except subprocess.CalledProcessError as e:
                raise GitCloneError(e.returncode,'https://gitee.com/zhtianyu/ycsb.git',e.stderr.decode('utf-8'))
        try:
            subprocess.run(
                "mvn -pl site.ycsb:redis-binding -am clean package",
                cwd=self.path,
                shell=True,check=True,
                stdout=subprocess.DEVNULL,
                stderr=subprocess.PIPE,
            )
        except subprocess.CalledProcessError as e:
            raise DefaultError(f"mvn命令运行失败,报错信息:{e.stderr.decode('utf-8')}")


        # 修改配置文件加入redis，重复运行时也避免配置项不断追加
        self._save_workload_config({
            'redis.host': '127.0.0.1',
            'redis.port': '6379',
        })

        # 加载数据
        self.load_result = self._run_ycsb_command(
            self.load_command,
            'ycsb_load.log',
            'ycsb load redis加载数据失败'
        )


    def run_test(self):
        self.test_result = self._run_ycsb_command(
            self.run_command,
            'ycsb.log',
            'ycsb run redis执行测试失败'
        )


    def result2summary(self):
        records = self._parse_records()
        grouped_records = self._group_records(records)
        return_metric_names = self._collect_return_metric_names(grouped_records)
        operation_rows = self._build_operation_rows(grouped_records, return_metric_names)
        gc_rows = self._build_gc_rows(grouped_records)
        workload_config = self._load_workload_config()
        command_args = self._parse_command_args(self.run_command)

        wb = Workbook()
        ws_summary = wb.active
        ws_summary.title = 'summary'
        ws_summary.append(['metric', 'value', 'note'])

        command_line_match = re.search(r"^Command line:\s*(.+)$", self.test_result, flags=re.MULTILINE)
        overall_metrics = grouped_records.get('OVERALL', {})
        total_gc = gc_rows[0]
        total_operations = sum(row['operations'] for row in operation_rows if isinstance(row['operations'], int))
        total_return_counts = {}
        for metric_name in return_metric_names:
            total_return_counts[metric_name] = sum(
                row['return_metrics'].get(metric_name, 0)
                for row in operation_rows
                if isinstance(row['return_metrics'].get(metric_name, ''), int)
            )
        total_return_ok = total_return_counts.get('Return=OK', 0)
        total_non_ok_returns = sum(
            value for metric_name, value in total_return_counts.items()
            if metric_name != 'Return=OK' and isinstance(value, int)
        )
        highest_avg_latency = max(operation_rows, key=lambda row: row['avg_latency_us'] if isinstance(row['avg_latency_us'], (int, float)) else float('-inf'))
        highest_p99_latency = max(operation_rows, key=lambda row: row['p99_latency_us'] if isinstance(row['p99_latency_us'], (int, float)) else float('-inf'))
        largest_operation = max(operation_rows, key=lambda row: row['operations'] if isinstance(row['operations'], int) else -1)
        configured_threads = command_args.get('-threads', '')
        overall_success_rate = ''
        if total_operations and isinstance(total_return_ok, int):
            overall_success_rate = round(total_return_ok / total_operations, 4)

        summary_rows = [
            ['run command', self.run_command, 'YCSB执行测试命令'],
            ['load command', self.load_command, 'YCSB加载数据命令'],
            ['java command line', command_line_match.group(1) if command_line_match else '', 'YCSB输出中的实际Java命令'],
            ['workload file', command_args.get('-P', ''), 'run阶段使用的workload配置文件'],
            ['configured threads', int(configured_threads) if str(configured_threads).isdigit() else configured_threads, 'YCSB并发线程数'],
            ['configured target ops/sec', command_args.get('-target', ''), '若为空表示未限制目标速率'],
            ['recordcount', workload_config.get('recordcount', ''), 'load阶段预加载记录数'],
            ['operationcount', workload_config.get('operationcount', ''), 'run阶段总操作数'],
            ['request distribution', workload_config.get('requestdistribution', ''), '请求分布模式'],
            ['read proportion', workload_config.get('readproportion', ''), '读操作占比'],
            ['update proportion', workload_config.get('updateproportion', ''), '更新操作占比'],
            ['scan proportion', workload_config.get('scanproportion', ''), '扫描操作占比'],
            ['insert proportion', workload_config.get('insertproportion', ''), '插入操作占比'],
            ['runtime(ms)', overall_metrics.get('RunTime(ms)', ''), '总运行时间'],
            ['throughput(ops/sec)', overall_metrics.get('Throughput(ops/sec)', ''), '越大越好'],
            ['total gc count', total_gc['count'], 'YCSB统计的GC总次数'],
            ['total gc time(ms)', total_gc['time_ms'], 'YCSB统计的GC总耗时'],
            ['total gc time(%)', total_gc['time_percent'], 'GC耗时占总时长百分比'],
            ['operation section count', len(operation_rows), '本次测试包含的操作类别数量'],
            ['total recorded operations', total_operations, '各操作的Operations求和'],
            ['return metric categories', ', '.join(return_metric_names), 'YCSB输出中出现的Return=XXX类别'],
            ['total return ok', total_return_ok, '各操作Return=OK求和'],
            ['total non-ok returns', total_non_ok_returns, '各操作非Return=OK求和'],
            ['overall success rate', self._format_ratio_percent(overall_success_rate), '总成功操作数/总操作数'],
            ['dominant operation', largest_operation['operation'], f"{largest_operation['operations']} ops"],
            ['highest avg latency operation', highest_avg_latency['operation'], f"{highest_avg_latency['avg_latency_us']} us"],
            ['highest p99 latency operation', highest_p99_latency['operation'], f"{highest_p99_latency['p99_latency_us']} us"],
        ]
        for metric_name in return_metric_names:
            summary_rows.append([
                f"total {metric_name}",
                total_return_counts.get(metric_name, ''),
                f"各操作{metric_name}求和",
            ])
        for row in summary_rows:
            ws_summary.append(row)

        ws_operations = wb.create_sheet(title='operations')
        operation_headers = ['operation', 'operations', 'operation ratio']
        operation_headers.extend(return_metric_names)
        operation_headers.extend([
            'non-ok return total',
            'success rate',
            'avg latency(us)', 'min latency(us)', 'max latency(us)',
            'p50 latency(us)', 'p95 latency(us)', 'p99 latency(us)',
        ])
        ws_operations.append(operation_headers)
        for row in operation_rows:
            operation_ratio = ''
            if total_operations and isinstance(row['operations'], int):
                operation_ratio = round(row['operations'] / total_operations, 4)
            ws_row = [
                row['operation'],
                row['operations'],
                self._format_ratio_percent(operation_ratio),
            ]
            ws_row.extend(row['return_metrics'].get(metric_name, '') for metric_name in return_metric_names)
            ws_row.extend([
                row['non_ok_return_total'],
                self._format_ratio_percent(row['success_rate']),
                row['avg_latency_us'],
                row['min_latency_us'],
                row['max_latency_us'],
                row['p50_latency_us'],
                row['p95_latency_us'],
                row['p99_latency_us'],
            ])
            ws_operations.append(ws_row)

        ws_returns = wb.create_sheet(title='return_codes')
        ws_returns.append(['operation', 'return metric', 'count', 'ratio in operation'])
        return_row_count = 0
        for row in operation_rows:
            operations = row['operations'] if isinstance(row['operations'], int) else 0
            for metric_name in return_metric_names:
                count = row['return_metrics'].get(metric_name, '')
                ratio = ''
                if operations and isinstance(count, int):
                    ratio = round(count / operations, 4)
                ws_returns.append([row['operation'], metric_name, count, self._format_ratio_percent(ratio)])
                return_row_count += 1
        if return_row_count == 0:
            ws_returns.append(['', '未提取到Return=XXX指标', '', ''])

        ws_gc = wb.create_sheet(title='gc')
        ws_gc.append(['collector', 'count', 'time(ms)', 'time(%)'])
        for row in gc_rows:
            ws_gc.append([row['collector'], row['count'], row['time_ms'], row['time_percent']])

        ws_config = wb.create_sheet(title='workload_config')
        ws_config.append(['key', 'value'])
        for key in sorted(workload_config):
            ws_config.append([key, workload_config[key]])

        ws_raw = wb.create_sheet(title='raw_metrics')
        ws_raw.append(['section', 'metric', 'value'])
        for record in records:
            ws_raw.append([record['section'], record['metric'], record['value']])

        for sheet in (ws_summary, ws_operations, ws_returns, ws_gc, ws_config, ws_raw):
            sheet.freeze_panes = 'A2'
            for row in sheet.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(vertical='top', wrap_text=True)

        ws_summary.column_dimensions['A'].width = 26
        ws_summary.column_dimensions['B'].width = 20
        ws_summary.column_dimensions['C'].width = 28

        for column_index in range(1, len(operation_headers) + 1):
            ws_operations.column_dimensions[get_column_letter(column_index)].width = 18

        for column in ('A', 'B', 'C', 'D'):
            ws_returns.column_dimensions[column].width = 20

        for column in ('A', 'B', 'C', 'D'):
            ws_gc.column_dimensions[column].width = 18

        ws_config.column_dimensions['A'].width = 24
        ws_config.column_dimensions['B'].width = 24

        ws_raw.column_dimensions['A'].width = 28
        ws_raw.column_dimensions['B'].width = 26
        ws_raw.column_dimensions['C'].width = 24

        wb.save(self.directory / 'ycsb.xlsx')



    def post_test(self):
        self.redis.Unit.Stop(b'replace')
        subprocess.run(
            "dnf remove -y redis",shell=True,stdout=subprocess.DEVNULL,stderr=subprocess.DEVNULL
        )


    def run(self):
        print("开始进行ycsb测试")
        self.pre_test()
        self.run_test()
        try:
            self.result2summary()
        except Exception as e:
            logFile = self.directory / 'ycsb_summary_error.log'
            with open(logFile,'w') as log:
                log.write(str(e))
            raise SummaryError(logFile)
        self.post_test()
        print("ycsb测试结束")

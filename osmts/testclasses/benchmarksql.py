from openpyxl.workbook import Workbook
from openpyxl.styles import Alignment
from pystemd.systemd1 import Unit
from pathlib import Path
from datetime import datetime
import re
import shlex
import time
import os
import requests,tarfile,zipfile
import pymysql
import subprocess,shutil
from io import BytesIO

from .errors import DefaultError,RunError,SummaryError


headers = {
    'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64; rv:137.0) Gecko/20100101 Firefox/137.0',
    'referer': 'https://gitee.com/April_Zhao/osmts',
}


class BenchMarkSQL:
    def __init__(self, **kwargs):
        self.rpms = {'postgresql-server','mysql-server','java'}
        self.directory: Path = kwargs.get('saved_directory') / 'benchmarksql'
        self.path:Path = Path('/root/osmts_tmp/benchmarksql')
        self.mysql_path: Path = self.path / 'mysql'
        self.postgresql_path: Path = self.path / 'postgresql'
        self.postgresql_hba_path = Path('/var/lib/pgsql/data/pg_hba.conf')

        self.mysql_build_result:str = ''
        self.mysql_test_result:str = ''
        self.postgres_build_result:str = ''
        self.postgres_test_result:str = ''
        self.mysql_build_info: dict = {}
        self.postgres_build_info: dict = {}
        self.mysql_artifacts: dict = {}
        self.postgres_artifacts: dict = {}
        self.java_heap_mb = int(kwargs.get('benchmarksql_java_heap_mb', 512))

        self.mysql_settings = {
            'warehouses': int(kwargs.get('benchmarksql_mysql_warehouses', 5)),
            'loadWorkers': int(kwargs.get('benchmarksql_mysql_load_workers', 4)),
            'terminals': int(kwargs.get('benchmarksql_mysql_terminals', 2)),
            'runTxnsPerTerminal': int(kwargs.get('benchmarksql_mysql_run_txns_per_terminal', 0)),
            'runMins': int(kwargs.get('benchmarksql_mysql_run_mins', 10)),
            'limitTxnsPerMin': int(kwargs.get('benchmarksql_mysql_limit_txns_per_min', 1000000000)),
        }
        self.postgres_settings = {
            'warehouses': int(kwargs.get('benchmarksql_postgres_warehouses', 1)),
            'loadWorkers': int(kwargs.get('benchmarksql_postgres_load_workers', 4)),
            'terminals': int(kwargs.get('benchmarksql_postgres_terminals', 1)),
            'runTxnsPerTerminal': int(kwargs.get('benchmarksql_postgres_run_txns_per_terminal', 10)),
            'runMins': int(kwargs.get('benchmarksql_postgres_run_mins', 0)),
            'limitTxnsPerMin': int(kwargs.get('benchmarksql_postgres_limit_txns_per_min', 300)),
        }


    # 下载指定 URL 的内容并返回字节数据
    def _download(self, url: str) -> bytes:
        response = requests.get(url=url, headers=headers, timeout=600)
        response.raise_for_status()
        return response.content


    # 下载 benchmarksql 压缩包并解压，分别复制为 MySQL 和 PostgreSQL 两份源码目录
    def _prepare_benchmarksql_sources(self):
        archive_content = self._download(
            "https://mirrors.huaweicloud.com/kunpeng/archive/kunpeng_solution/database/patch/benchmarksql5.0-for-mysql.zip"
        )

        extract_root = self.path / '_archive'
        shutil.rmtree(extract_root, ignore_errors=True)
        extract_root.mkdir(parents=True, exist_ok=True)

        try:
            with zipfile.ZipFile(BytesIO(archive_content)) as zip_file:
                zip_file.extractall(extract_root)
        except zipfile.BadZipFile as e:
            raise DefaultError(f"benchmarksql测试出错.下载的benchmarksql压缩包不是有效zip文件,报错信息:{str(e)}")

        candidate_dirs = [path for path in extract_root.iterdir() if path.is_dir()]
        if len(candidate_dirs) == 1:
            source_dir = candidate_dirs[0]
        else:
            source_dir = extract_root / 'benchmarksql5.0-for-mysql'

        if not source_dir.exists():
            raise DefaultError(f"benchmarksql测试出错.解压后未找到benchmarksql目录:{source_dir}")

        shutil.rmtree(self.mysql_path, ignore_errors=True)
        shutil.rmtree(self.postgresql_path, ignore_errors=True)
        shutil.copytree(source_dir, self.mysql_path)
        shutil.copytree(source_dir, self.postgresql_path)


    # 在文本中查找 pattern 的最后一个匹配，返回第一个捕获组（无匹配返回空字符串）
    def _find_last_group(self, text: str, pattern: str):
        last_match = None
        for match in re.finditer(pattern, text, flags=re.MULTILINE):
            last_match = match
        if last_match is None:
            return ''
        return last_match.group(1).strip()


    # 将单次测试运行的关键参数（仓库数、终端数等）格式化为可读备注字符串
    def _format_run_note(self, run: dict):
        fallback_parts = []
        if run.get('phase'):
            fallback_parts.append(f"phase={run['phase']}")
        if run.get('load_workers') != '':
            fallback_parts.append(f"loadWorkers={run['load_workers']}")
        if run['run_type'] and run['run_value'] != '':
            fallback_parts.append(f"{run['run_type']}={run['run_value']}")
        if run['warehouses'] != '':
            fallback_parts.append(f"warehouses={run['warehouses']}")
        if run['terminals'] != '':
            fallback_parts.append(f"terminals={run['terminals']}")
        return ', '.join(fallback_parts)


    # 规范化数据库名称，将 "postgres" 统一映射为 "postgresql"
    def _normalize_db_name(self, db_name: str):
        normalized = (db_name or '').strip().lower()
        if normalized == 'postgres':
            return 'postgresql'
        return normalized


    # 清理日志文本中的 \r 和退格符，统一换行格式
    def _clean_log_text(self, text: str):
        return text.replace('\r', '\n').replace('\x08', '')


    # 按 warmup/start test 阶段标记将日志文本切分为多个阶段块
    def _split_phase_blocks(self, text: str):
        clean_text = self._clean_log_text(text)
        phase_pattern = re.compile(
            r"[A-Z][a-z]{2} [A-Z][a-z]{2}\s+\d{1,2} "
            r"\d{2}:\d{2}:\d{2} [AP]M [A-Z]{3} \d{4} -- (?P<phase>warmup|start test)"
        )
        matches = list(phase_pattern.finditer(clean_text))
        if not matches:
            return [{'phase': '', 'content': clean_text}]

        blocks = []
        for idx, match in enumerate(matches):
            end = matches[idx + 1].start() if idx + 1 < len(matches) else len(clean_text)
            blocks.append({
                'phase': match.group('phase'),
                'content': clean_text[match.start():end],
            })
        return blocks


    # 计算会话起止时间之差（秒），解析失败时返回空字符串
    def _safe_session_duration(self, session_start: str, session_end: str):
        try:
            start_dt = datetime.strptime(session_start, '%Y-%m-%d %H:%M:%S')
            end_dt = datetime.strptime(session_end, '%Y-%m-%d %H:%M:%S')
        except ValueError:
            return ''
        return int((end_dt - start_dt).total_seconds())


    # 从日志中提取结果目录路径并将 run.properties/runInfo.csv/result.csv 复制到保存目录
    def _copy_result_artifacts(self, run_root: Path, content: str, db_name: str):
        clean_text = self._clean_log_text(content)
        artifacts: dict[str, dict] = {}

        def ensure_entry(dir_name: str):
            return artifacts.setdefault(dir_name, {
                'result_dir': dir_name,
                'run_properties_rel': '',
                'runinfo_rel': '',
                'result_csv_rel': '',
                'run_id': '',
                'saved_dir': '',
                'saved_run_properties': '',
                'saved_runinfo_csv': '',
                'saved_result_csv': '',
            })

        for match in re.finditer(r'copied\s+\S+\s+to\s+(\S+)/run\.properties', clean_text):
            result_dir = Path(match.group(1)).parts[0]
            entry = ensure_entry(result_dir)
            entry['run_properties_rel'] = f"{result_dir}/run.properties"

        for match in re.finditer(r'created\s+(\S+/runInfo\.csv)\s+for runID\s+(\d+)', clean_text):
            runinfo_rel = match.group(1)
            result_dir = Path(runinfo_rel).parts[0]
            entry = ensure_entry(result_dir)
            entry['runinfo_rel'] = runinfo_rel
            entry['run_id'] = int(match.group(2))

        for match in re.finditer(r'writing per transaction results to\s+(\S+/result\.csv)', clean_text):
            result_csv_rel = match.group(1)
            result_dir = Path(result_csv_rel).parts[0]
            entry = ensure_entry(result_dir)
            entry['result_csv_rel'] = result_csv_rel

        saved_root = self.directory / f'{db_name}_artifacts'
        saved_root.mkdir(parents=True, exist_ok=True)
        for result_dir, entry in artifacts.items():
            source_dir = run_root / result_dir
            if not source_dir.exists():
                continue

            target_dir = saved_root / result_dir
            if target_dir.exists():
                shutil.rmtree(target_dir)
            shutil.copytree(source_dir, target_dir)
            entry['saved_dir'] = target_dir

            for rel_key, saved_key in (
                ('run_properties_rel', 'saved_run_properties'),
                ('runinfo_rel', 'saved_runinfo_csv'),
                ('result_csv_rel', 'saved_result_csv'),
            ):
                rel_path = entry.get(rel_key, '')
                if not rel_path:
                    continue
                rel_subpath = Path(rel_path).relative_to(result_dir)
                saved_path = target_dir / rel_subpath
                if saved_path.exists():
                    entry[saved_key] = saved_path

        return artifacts


    # 从 build 阶段日志中提取数据库配置参数和数据加载完成情况
    def _parse_build_info(self, build_result: str, default_db: str, log_path: Path):
        clean_text = self._clean_log_text(build_result)
        warehouses_text = self._find_last_group(clean_text, r"\bwarehouses\s*[:=]\s*(\d+)")
        load_workers_text = self._find_last_group(clean_text, r"\bloadWorkers\s*[:=]\s*(\d+)")
        loaded_warehouses = sorted(
            {int(item) for item in re.findall(r'Loading Warehouse\s+(\d+)\s+done', clean_text)}
        )
        return {
            'db': self._normalize_db_name(default_db),
            'driver': self._find_last_group(clean_text, r"\bdriver\s*[:=]\s*([^\n\r]+)"),
            'conn': self._find_last_group(clean_text, r"\bconn\s*[:=]\s*([^\n\r]+)"),
            'user': self._find_last_group(clean_text, r"\buser\s*[:=]\s*([^\n\r]+)"),
            'warehouses': int(warehouses_text) if warehouses_text else '',
            'load_workers': int(load_workers_text) if load_workers_text else '',
            'file_location': self._find_last_group(clean_text, r"\bfileLocation\b\s*([^\n\r]+)"),
            'csv_null_value': self._find_last_group(clean_text, r"\bcsvNullValue\b\s*([^\n\r]+)"),
            'item_loaded': 'yes' if 'Loading ITEM done' in clean_text else 'no',
            'loaded_warehouse_count': len(loaded_warehouses),
            'loaded_warehouses': ', '.join(str(item) for item in loaded_warehouses),
            'index_creates': 'yes' if 'indexCreates.sql' in clean_text else 'no',
            'foreign_keys': 'yes' if 'foreignKeys.sql' in clean_text else 'no',
            'build_finish': 'yes' if 'buildFinish.sql' in clean_text else 'no',
            'log_path': log_path,
        }


    # 读取 MySQL tpcc.bmsql_config 表内容，返回 cfg_name -> cfg_value 映射
    def _read_mysql_bmsql_config(self):
        try:
            conn = pymysql.connect(
                host='localhost',
                port=3306,
                user='root',
                passwd='123456',
                database='tpcc',
            )
        except Exception as e:
            raise DefaultError(
                "benchmarksql测试出错.连接MySQL tpcc数据库读取bmsql_config失败,报错信息:"
                + str(e)
            )

        try:
            cursor = conn.cursor()
            cursor.execute("SHOW TABLES LIKE 'bmsql_config';")
            if cursor.fetchone() is None:
                raise DefaultError("benchmarksql测试出错.MySQL建库后缺少tpcc.bmsql_config表.")

            cursor.execute("SELECT cfg_name, cfg_value FROM bmsql_config ORDER BY cfg_name;")
            rows = cursor.fetchall()
        except Exception as e:
            raise DefaultError(
                "benchmarksql测试出错.读取MySQL bmsql_config表失败,报错信息:"
                + str(e)
            )
        finally:
            cursor.close()
            conn.close()

        return {str(name): str(value) for name, value in rows}


    # 保存 MySQL bmsql_config 快照，便于后续排查 build 阶段问题
    def _save_mysql_bmsql_config_snapshot(self, config_rows: dict[str, str]):
        snapshot_path = self.directory / 'mysql_bmsql_config.log'
        with open(snapshot_path, 'w', encoding='utf-8') as log:
            if not config_rows:
                log.write('<empty>\n')
            else:
                for key, value in config_rows.items():
                    log.write(f'{key}={value}\n')
        return snapshot_path


    # 校验 MySQL build 阶段是否真正完成，避免 runDatabaseBuild.sh 退出码为0但数据库装载并未成功
    def _validate_mysql_build_state(self, build_info: dict, log_path: Path):
        expected_warehouses = self.mysql_settings['warehouses']
        problems = []

        if build_info.get('item_loaded') != 'yes':
            problems.append('mysql_build.log中未检测到Loading ITEM done')
        if build_info.get('loaded_warehouse_count', 0) < expected_warehouses:
            problems.append(
                f"mysql_build.log中完成加载的仓库数不足, 期望:{expected_warehouses}, 实际:{build_info.get('loaded_warehouse_count', 0)}"
            )
        if build_info.get('foreign_keys') != 'yes':
            problems.append('mysql_build.log中未检测到foreignKeys.sql执行完成')
        if build_info.get('build_finish') != 'yes':
            problems.append('mysql_build.log中未检测到buildFinish.sql执行完成')

        config_rows = self._read_mysql_bmsql_config()
        snapshot_path = self._save_mysql_bmsql_config_snapshot(config_rows)
        build_info['config_row_count'] = len(config_rows)
        build_info['config_warehouses'] = config_rows.get('warehouses', '')
        build_info['config_snapshot_path'] = snapshot_path

        config_warehouses = config_rows.get('warehouses', '')
        if not config_warehouses:
            problems.append(f"tpcc.bmsql_config中缺少warehouses项, 快照请查看:{snapshot_path}")
        elif str(config_warehouses) != str(expected_warehouses):
            problems.append(
                f"tpcc.bmsql_config中的warehouses值异常, 期望:{expected_warehouses}, 实际:{config_warehouses}, 快照请查看:{snapshot_path}"
            )

        if problems:
            raise DefaultError(
                "benchmarksql测试出错.MySQL建库/装数阶段校验失败:"
                + '；'.join(problems)
                + f". 详细信息请查看:{log_path}"
            )


    # 从测试阶段日志中提取每次运行的 tpmC、tpmTOTAL、会话时长等性能指标
    def _parse_runs(self, test_result: str, default_db: str, build_info: dict | None = None, artifact_map: dict | None = None):
        result_pattern = re.compile(
            r"Measured tpmC \(NewOrders\)\s*=\s*([\d.]+).*?"
            r"Measured tpmTOTAL\s*=\s*([\d.]+).*?"
            r"Session Start\s*=\s*(.+?)\s*\n.*?"
            r"Session End\s*=\s*(.+?)\s*\n.*?"
            r"Transaction Count\s*=\s*(\d+)",
            flags=re.DOTALL,
        )
        build_info = build_info or {}
        artifact_map = artifact_map or {}
        clean_text = self._clean_log_text(test_result)
        java_options = self._find_last_group(clean_text, r"_JAVA_OPTIONS:\s*([^\n\r]+)")
        runs = []

        for block in self._split_phase_blocks(clean_text):
            matches = list(result_pattern.finditer(block['content']))
            if not matches:
                continue

            for match in matches:
                context = block['content'][:match.start()]
                db_name = self._normalize_db_name(
                    self._find_last_group(context, r"\bdb\s*[:=]\s*([^\s,]+)") or default_db
                )
                warehouses_text = self._find_last_group(context, r"\bwarehouses\s*[:=]\s*(\d+)")
                terminals_text = self._find_last_group(context, r"\bterminals\s*[:=]\s*(\d+)")
                run_mins_text = self._find_last_group(context, r"\brunMins\s*[:=]\s*(\d+)")
                run_txns_text = self._find_last_group(context, r"\brunTxnsPerTerminal\s*[:=]\s*(\d+)")
                limit_txns_text = self._find_last_group(context, r"\blimitTxnsPerMin\s*[:=]\s*(\d+)")
                terminal_fixed = self._find_last_group(context, r"\bterminalWarehouseFixed\s*[:=]\s*([^\s,]+)")
                new_order_weight = self._find_last_group(context, r"\bnewOrderWeight\s*[:=]\s*(\d+)")
                payment_weight = self._find_last_group(context, r"\bpaymentWeight\s*[:=]\s*(\d+)")
                order_status_weight = self._find_last_group(context, r"\borderStatusWeight\s*[:=]\s*(\d+)")
                delivery_weight = self._find_last_group(context, r"\bdeliveryWeight\s*[:=]\s*(\d+)")
                stock_level_weight = self._find_last_group(context, r"\bstockLevelWeight\s*[:=]\s*(\d+)")
                c_last_load = self._find_last_group(context, r"C value for C_LAST during load:\s*(\d+)")
                c_last_run = self._find_last_group(context, r"C value for C_LAST this run:\s*(\d+)")
                runinfo_rel = self._find_last_group(context, r'created\s+(\S+/runInfo\.csv)\s+for runID\s+\d+')
                result_csv_rel = self._find_last_group(context, r'writing per transaction results to\s+(\S+/result\.csv)')
                run_properties_dir = self._find_last_group(context, r'copied\s+\S+\s+to\s+(\S+)/run\.properties')
                run_id_text = self._find_last_group(context, r'created\s+\S+/runInfo\.csv\s+for runID\s+(\d+)')

                result_dir = ''
                for candidate in (run_properties_dir, runinfo_rel, result_csv_rel):
                    if candidate:
                        result_dir = Path(candidate).parts[0]
                        break
                artifact_info = artifact_map.get(result_dir, {})

                run_type = ''
                run_value = ''
                if run_mins_text:
                    run_type = 'runMins'
                    run_value = int(run_mins_text)
                elif run_txns_text:
                    run_type = 'runTxnsPerTerminal'
                    run_value = int(run_txns_text)

                session_start = match.group(3).strip()
                session_end = match.group(4).strip()
                runs.append({
                    'db': db_name,
                    'phase': block['phase'],
                    'driver': self._find_last_group(context, r"\bdriver\s*[:=]\s*([^\n\r]+)"),
                    'conn': self._find_last_group(context, r"\bconn\s*[:=]\s*([^\n\r]+)"),
                    'user': self._find_last_group(context, r"\buser\s*[:=]\s*([^\n\r]+)"),
                    'warehouses': int(warehouses_text) if warehouses_text else build_info.get('warehouses', ''),
                    'load_workers': build_info.get('load_workers', ''),
                    'terminals': int(terminals_text) if terminals_text else '',
                    'run_type': run_type,
                    'run_value': run_value,
                    'limit_txns_per_min': int(limit_txns_text) if limit_txns_text else '',
                    'terminal_warehouse_fixed': terminal_fixed,
                    'new_order_weight': int(new_order_weight) if new_order_weight else '',
                    'payment_weight': int(payment_weight) if payment_weight else '',
                    'order_status_weight': int(order_status_weight) if order_status_weight else '',
                    'delivery_weight': int(delivery_weight) if delivery_weight else '',
                    'stock_level_weight': int(stock_level_weight) if stock_level_weight else '',
                    'measured_tpmc': float(match.group(1)),
                    'measured_tpmtotal': float(match.group(2)),
                    'transaction_count': int(match.group(5)),
                    'session_start': session_start,
                    'session_end': session_end,
                    'session_duration_seconds': self._safe_session_duration(session_start, session_end),
                    'c_last_during_load': int(c_last_load) if c_last_load else '',
                    'c_last_this_run': int(c_last_run) if c_last_run else '',
                    'run_id': artifact_info.get('run_id', int(run_id_text) if run_id_text else ''),
                    'result_dir': result_dir,
                    'saved_artifact_dir': str(artifact_info.get('saved_dir', '')) if artifact_info.get('saved_dir') else '',
                    'saved_run_properties': str(artifact_info.get('saved_run_properties', '')) if artifact_info.get('saved_run_properties') else '',
                    'saved_runinfo_csv': str(artifact_info.get('saved_runinfo_csv', '')) if artifact_info.get('saved_runinfo_csv') else '',
                    'saved_result_csv': str(artifact_info.get('saved_result_csv', '')) if artifact_info.get('saved_result_csv') else '',
                    'java_options': java_options,
                })

        if not runs:
            raise ValueError('benchmarksql日志中未找到可解析的测试轮次')
        return runs


    # 向工作表写入每次测试运行的完整指标行（tpmC、tpmTOTAL、配置参数等）
    def _populate_db_sheet(self, ws, runs):
        ws.append([
            'db', 'phase', 'driver', 'conn', 'user',
            'warehouses', 'loadWorkers', 'terminals', 'run type', 'run value',
            'limitTxnsPerMin', 'terminalWarehouseFixed',
            'newOrderWeight', 'paymentWeight', 'orderStatusWeight', 'deliveryWeight', 'stockLevelWeight',
            'measured tpmC', 'measured tpmTOTAL', 'transaction count',
            'session start', 'session end', 'session duration(s)',
            'C_LAST during load', 'C_LAST this run', 'run id',
            'result directory', 'saved artifact dir', 'saved run.properties', 'saved runInfo.csv', 'saved result.csv',
            'java options',
        ])
        for run in runs:
            ws.append([
                run['db'],
                run['phase'],
                run['driver'],
                run['conn'],
                run['user'],
                run['warehouses'],
                run['load_workers'],
                run['terminals'],
                run['run_type'],
                run['run_value'],
                run['limit_txns_per_min'],
                run['terminal_warehouse_fixed'],
                run['new_order_weight'],
                run['payment_weight'],
                run['order_status_weight'],
                run['delivery_weight'],
                run['stock_level_weight'],
                run['measured_tpmc'],
                run['measured_tpmtotal'],
                run['transaction_count'],
                run['session_start'],
                run['session_end'],
                run['session_duration_seconds'],
                run['c_last_during_load'],
                run['c_last_this_run'],
                run['run_id'],
                run['result_dir'],
                run['saved_artifact_dir'],
                run['saved_run_properties'],
                run['saved_runinfo_csv'],
                run['saved_result_csv'],
                run['java_options'],
            ])


    # 向工作表写入数据库 build 阶段的配置信息和数据加载完成情况
    def _populate_build_sheet(self, ws, rows):
        ws.append([
            'db', 'driver', 'conn', 'user', 'warehouses', 'loadWorkers',
            'fileLocation', 'csvNullValue', 'item loaded', 'loaded warehouse count',
            'loaded warehouses', 'indexCreates', 'foreignKeys', 'buildFinish',
            'config row count', 'config warehouses', 'config snapshot', 'log file',
        ])
        for row in rows:
            ws.append([
                row['db'],
                row['driver'],
                row['conn'],
                row['user'],
                row['warehouses'],
                row['load_workers'],
                row['file_location'],
                row['csv_null_value'],
                row['item_loaded'],
                row['loaded_warehouse_count'],
                row['loaded_warehouses'],
                row['index_creates'],
                row['foreign_keys'],
                row['build_finish'],
                row.get('config_row_count', ''),
                row.get('config_warehouses', ''),
                str(row.get('config_snapshot_path', '')) if row.get('config_snapshot_path') else '',
                str(row['log_path']),
            ])


    # 优先返回 start test 阶段的正式运行记录，无正式记录则返回全部
    def _select_summary_runs(self, runs):
        formal_runs = [run for run in runs if run['phase'] == 'start test']
        return formal_runs if formal_runs else runs


    # 将指定目录下所有 .sh 文件设置为可执行权限
    def _ensure_shell_scripts_executable(self, run_path: Path):
        if not run_path.exists():
            raise DefaultError(f"benchmarksql测试出错.目录不存在:{run_path}")
        for shell_file in run_path.glob('*.sh'):
            shell_file.chmod(0o755)


    # 将指定键值覆写到 properties 配置文件，保留其余行不变
    def _update_properties_file(self, file_path: Path, overrides: dict[str, str]):
        if not file_path.exists():
            raise DefaultError(f"benchmarksql测试出错.缺少配置文件:{file_path}")

        with open(file_path, 'r', encoding='utf-8') as file:
            original_lines = file.readlines()

        updated_lines = []
        updated_keys = set()
        for raw_line in original_lines:
            stripped = raw_line.strip()
            if not stripped or stripped.startswith('#') or '=' not in raw_line:
                updated_lines.append(raw_line)
                continue

            key = raw_line.split('=', 1)[0].strip()
            if key in overrides:
                line_ending = '\n' if raw_line.endswith('\n') else ''
                updated_lines.append(f"{key}={overrides[key]}{line_ending}")
                updated_keys.add(key)
            else:
                updated_lines.append(raw_line)

        if updated_lines and not updated_lines[-1].endswith('\n'):
            updated_lines[-1] += '\n'

        for key, value in overrides.items():
            if key not in updated_keys:
                updated_lines.append(f"{key}={value}\n")

        with open(file_path, 'w', encoding='utf-8') as file:
            file.writelines(updated_lines)


    # 确保 properties 文件存在（不存在时从模板复制），并写入指定覆盖参数
    def _ensure_properties_file(self, run_path: Path, properties_name: str, overrides: dict[str, str]):
        properties_path = run_path / properties_name
        template_path = run_path / 'props.conf'
        if not properties_path.exists():
            if not template_path.exists():
                raise DefaultError(f"benchmarksql测试出错.缺少模板配置文件:{template_path}")
            shutil.copy2(template_path, properties_path)
        self._update_properties_file(properties_path, overrides)


    # 生成数据库配置字典（MySQL 或 PostgreSQL）
    def _get_db_config(self, db_type: str) -> dict[str, str]:
        settings = self.mysql_settings if db_type == 'mysql' else self.postgres_settings
        base_config = {
            'warehouses': str(settings['warehouses']),
            'loadWorkers': str(settings['loadWorkers']),
            'terminals': str(settings['terminals']),
            'runTxnsPerTerminal': str(settings['runTxnsPerTerminal']),
            'runMins': str(settings['runMins']),
            'limitTxnsPerMin': str(settings['limitTxnsPerMin']),
            'terminalWarehouseFixed': 'true',
            'user': 'root' if db_type == 'mysql' else 'postgres',
            'password': '123456',
        }

        if db_type == 'mysql':
            base_config.update({
                'db': 'mysql',
                'driver': 'com.mysql.cj.jdbc.Driver',
                'conn': 'jdbc:mysql://localhost:3306/tpcc?allowPublicKeyRetrieval=true&useSSL=false&useServerPrepStmts=true&useConfigs=maxPerformance&rewriteBatchedStatements=true',
            })
        else:
            base_config.update({
                'db': 'postgres',
                'driver': 'org.postgresql.Driver',
                'conn': 'jdbc:postgresql://localhost:5432/tpcc',
            })

        return base_config


    # 修改 PostgreSQL build 脚本，确保 AFTER_LOAD 包含完整的索引和外键创建步骤
    def _patch_postgresql_build_script(self):
        build_script = self.postgresql_path / 'run' / 'runDatabaseBuild.sh'
        if not build_script.exists():
            raise DefaultError(f"benchmarksql测试出错.缺少脚本文件:{build_script}")

        with open(build_script, 'r', encoding='utf-8') as file:
            content = file.read()

        patched = re.sub(
            r'(?m)^#?AFTER_LOAD="indexCreates foreignKeys buildFinish"\s*$',
            'AFTER_LOAD="indexCreates foreignKeys buildFinish"',
            content,
        )
        patched = re.sub(
            r'(?m)^#?AFTER_LOAD="foreignKeys buildFinish"\s*$',
            '#AFTER_LOAD="foreignKeys buildFinish"',
            patched,
        )

        if patched != content:
            with open(build_script, 'w', encoding='utf-8') as file:
                file.write(patched)


    # 检查命令是否在 PATH 中可用，不可用则抛出含安装提示的错误
    def _ensure_command_available(self, command_name: str, install_hint: str):
        command_path = shutil.which(command_name)
        if command_path:
            return command_path
        raise DefaultError(
            f"benchmarksql测试出错.未找到{command_name}命令.当前测试依赖该客户端工具."
            f"请确认{install_hint}已正确安装; 如果当前系统将客户端单独打包,请额外安装对应客户端包后重试."
        )


    # 以 postgres 用户身份执行单条 SQL，失败时抛出含错误信息的异常
    def _run_postgres_psql(self, sql: str):
        psql_path = self._ensure_command_available('psql', 'postgresql-server')
        try:
            subprocess.run(
                [
                    'su', '-', 'postgres', '-c',
                    f"{shlex.quote(psql_path)} -v ON_ERROR_STOP=1 -d postgres -c {shlex.quote(sql)}",
                ],
                check=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
            )
        except subprocess.CalledProcessError as e:
            raise DefaultError(
                "benchmarksql测试出错.PostgreSQL初始化SQL执行失败,报错信息:"
                + e.stderr.decode('utf-8', errors='replace')
            )


    # 重启 postgresql.service 并等待其回到 active 状态
    def _restart_postgresql_service(self):
        try:
            subprocess.run(
                "systemctl restart postgresql.service",
                shell=True,
                check=True,
                stdout=subprocess.DEVNULL,
                stderr=subprocess.PIPE,
            )
        except subprocess.CalledProcessError as e:
            raise DefaultError(f"benchmarksql测试出错.重启postgresql.service失败,报错信息:{e.stderr.decode('utf-8')}")

        time.sleep(5)
        self.postgresql.load(force=True)
        if self.postgresql.Unit.ActiveState != b'active':
            raise DefaultError("benchmarksql测试出错.重启postgresql.service后服务未处于active状态.")


    # 确保 pg_hba.conf 中本地连接使用 trust 认证，有改动时重启服务
    def _ensure_postgresql_local_auth(self):
        if not self.postgresql_hba_path.exists():
            return

        changed = False
        updated_lines = []
        for raw_line in self.postgresql_hba_path.read_text(encoding='utf-8').splitlines(keepends=True):
            stripped = raw_line.strip()
            if stripped.startswith('host') and '127.0.0.1/32' in stripped and not stripped.startswith('#'):
                parts = raw_line.split()
                if parts and parts[-1] != 'trust':
                    parts[-1] = 'trust'
                    raw_line = '    '.join(parts) + '\n'
                    changed = True
            elif stripped.startswith('host') and '::1/128' in stripped and not stripped.startswith('#'):
                parts = raw_line.split()
                if parts and parts[-1] != 'trust':
                    parts[-1] = 'trust'
                    raw_line = '    '.join(parts) + '\n'
                    changed = True
            updated_lines.append(raw_line)

        if changed:
            with open(self.postgresql_hba_path, 'w', encoding='utf-8') as file:
                file.writelines(updated_lines)
            self._restart_postgresql_service()


    # 构造 benchmarksql 运行所需的环境变量，设置 JVM 堆大小
    def _benchmarksql_env(self):
        env = os.environ.copy()
        env['_JAVA_OPTIONS'] = f"-Xms128m -Xmx{self.java_heap_mb}m"
        return env


    # 将进程的 stdout/stderr 写入日志文件并返回合并后的文本内容
    def _save_process_output(self, completed_process: subprocess.CompletedProcess, file_path: Path) -> str:
        stdout = completed_process.stdout.decode('utf-8', errors='replace')
        stderr = completed_process.stderr.decode('utf-8', errors='replace')
        content = stdout
        if stderr.strip():
            content = f"{stdout}\n\n[stderr]\n{stderr}".strip() + '\n'
        with open(file_path, 'w', encoding='utf-8') as file:
            file.write(content)
        return content


    # 检测日志中的已知错误模式（OOM、配置缺失等），有则抛出含原因的异常
    def _raise_if_known_failure(self, content: str, log_path: Path, stage_desc: str):
        known_patterns = [
            ('OutOfMemoryError', 'Java堆内存不足'),
            ("DB Load configuration parameter 'warehouses' not found", '数据库装载配置缺失，通常说明runDatabaseBuild阶段未正确完成'),
        ]
        for pattern, reason in known_patterns:
            if pattern in content:
                raise DefaultError(
                    f"benchmarksql测试出错.{stage_desc}检测到已知错误:{reason}.详细信息请查看:{log_path}"
                )


    # 执行 benchmarksql 脚本（build 或 benchmark），保存日志并检查错误
    def _run_benchmarksql_script(self, script_name: str, properties_name: str, work_dir: Path, log_file: Path, stage_desc: str) -> str:
        try:
            result = subprocess.run(
                f"./{script_name} {properties_name}",
                cwd=work_dir,
                shell=True, check=True,
                env=self._benchmarksql_env(),
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
            )
        except subprocess.CalledProcessError as e:
            raise RunError(e.returncode, e.stderr.decode('utf-8'))

        output = self._save_process_output(result, log_file)
        self._raise_if_known_failure(output, log_file, stage_desc)
        return output


    def pre_test(self):
        if self.directory.exists():
            shutil.rmtree(self.directory)
        self.directory.mkdir(parents=True)
        shutil.rmtree(self.path, ignore_errors=True)
        self.path.mkdir(parents=True, exist_ok=True)

        # 初始化MySQL
        self.mysqld:Unit = Unit('mysqld.service',_autoload=True)
        try:
            self.mysqld.Unit.Start(b'replace')
        except:
            time.sleep(5)
            self.mysqld.load(force=True)
            self.mysqld.Unit.Start(b'replace')
        time.sleep(5)
        if self.mysqld.Unit.ActiveState != b'active':
            time.sleep(5)
            if self.mysqld.Unit.ActiveState != b'active':
                raise DefaultError(f"benchmarksql测试出错.开启mysqld.service失败,退出测试.")

        try:
            self.mysql_conn = pymysql.connect(
                host='localhost',
                port=3306,
                user='root',
                passwd='',
            )
        except Exception as e:
            self.mysql_conn = pymysql.connect(
                host='localhost',
                port=3306,
                user='root',
                passwd='123456',
            )
        cursor = self.mysql_conn.cursor()
        cursor.execute("ALTER USER 'root'@'localhost' IDENTIFIED BY '123456';")
        cursor.execute("DROP DATABASE IF EXISTS tpcc;")
        cursor.execute("CREATE DATABASE IF NOT EXISTS tpcc;")
        self.mysql_conn.commit()
        cursor.close()
        self.mysql_conn.close()

        # 获取benchmark for mysql
        self._prepare_benchmarksql_sources()
        self._ensure_shell_scripts_executable(self.mysql_path / 'run')
        self._ensure_properties_file(
            self.mysql_path / 'run',
            'mysql.properties',
            self._get_db_config('mysql'),
        )

        # -------------------------------------------------------

        # 初始化postgresql
        if not self.postgresql_hba_path.exists():
            try:
                subprocess.run(
                    "/usr/bin/postgresql-setup initdb",
                    shell=True,check=True,
                    stdout=subprocess.DEVNULL,
                    stderr=subprocess.PIPE
                )
            except subprocess.CalledProcessError as e:
                raise DefaultError(f"benchmarksql测试出错.初始化postgresql数据库失败,报错信息:{e.stderr.decode('utf-8')}")

        self.postgresql:Unit = Unit('postgresql.service',_autoload=True)
        try:
            self.postgresql.Unit.Start(b'replace')
        except Exception:
            time.sleep(5)
            self.postgresql.load(force=True)
            self.postgresql.Unit.Start(b'replace')
        time.sleep(5)
        if self.postgresql.Unit.ActiveState != b'active':
            time.sleep(5)
            if self.postgresql.Unit.ActiveState != b'active':
                raise DefaultError("benchmarksql测试出错.开启postgresql.service失败.")

        self._ensure_postgresql_local_auth()
        self._run_postgres_psql("ALTER USER postgres WITH PASSWORD '123456';")
        self._run_postgres_psql("DROP DATABASE IF EXISTS tpcc;")
        self._run_postgres_psql("CREATE DATABASE tpcc;")

        self._ensure_shell_scripts_executable(self.postgresql_path / 'run')
        self._ensure_properties_file(
            self.postgresql_path / 'run',
            'postgres.properties',
            self._get_db_config('postgres'),
        )
        self._patch_postgresql_build_script()



    def run_test(self):
        # MySQL build
        self.mysql_build_result = self._run_benchmarksql_script(
            'runDatabaseBuild.sh', 'mysql.properties',
            self.mysql_path / 'run', self.directory / 'mysql_build.log',
            'MySQL建库/装数阶段'
        )
        self.mysql_build_info = self._parse_build_info(self.mysql_build_result, 'mysql', self.directory / 'mysql_build.log')
        self._validate_mysql_build_state(self.mysql_build_info, self.directory / 'mysql_build.log')

        # MySQL benchmark
        self.mysql_test_result = self._run_benchmarksql_script(
            'runBenchmark.sh', 'mysql.properties',
            self.mysql_path / 'run', self.directory / 'mysql.log',
            'MySQL压测阶段'
        )
        self.mysql_artifacts = self._copy_result_artifacts(self.mysql_path / 'run', self.mysql_test_result, 'mysql')

        # PostgreSQL build
        self.postgres_build_result = self._run_benchmarksql_script(
            'runDatabaseBuild.sh', 'postgres.properties',
            self.postgresql_path / 'run', self.directory / 'postgresql_build.log',
            'PostgreSQL建库/装数阶段'
        )
        self.postgres_build_info = self._parse_build_info(self.postgres_build_result, 'postgresql', self.directory / 'postgresql_build.log')

        # PostgreSQL benchmark
        self.postgres_test_result = self._run_benchmarksql_script(
            'runBenchmark.sh', 'postgres.properties',
            self.postgresql_path / 'run', self.directory / 'postgresql.log',
            'PostgreSQL压测阶段'
        )
        self.postgres_artifacts = self._copy_result_artifacts(self.postgresql_path / 'run', self.postgres_test_result, 'postgresql')


    def result2summary(self):
        mysql_runs = self._parse_runs(
            self.mysql_test_result,
            'mysql',
            build_info=self.mysql_build_info,
            artifact_map=self.mysql_artifacts,
        )
        postgres_runs = self._parse_runs(
            self.postgres_test_result,
            'postgresql',
            build_info=self.postgres_build_info,
            artifact_map=self.postgres_artifacts,
        )
        mysql_summary_runs = self._select_summary_runs(mysql_runs)
        postgres_summary_runs = self._select_summary_runs(postgres_runs)

        wb = Workbook()
        ws_summary = wb.active
        ws_summary.title = 'summary'
        ws_summary.append(['db', 'metric', 'value', 'note'])

        best_mysql_tpmc = max(mysql_summary_runs, key=lambda run: run['measured_tpmc'])
        best_mysql_tpmtotal = max(mysql_summary_runs, key=lambda run: run['measured_tpmtotal'])
        best_postgres_tpmc = max(postgres_summary_runs, key=lambda run: run['measured_tpmc'])
        best_postgres_tpmtotal = max(postgres_summary_runs, key=lambda run: run['measured_tpmtotal'])

        summary_rows = [
            ['mysql', 'best measured tpmC', best_mysql_tpmc['measured_tpmc'], self._format_run_note(best_mysql_tpmc)],
            ['mysql', 'best measured tpmTOTAL', best_mysql_tpmtotal['measured_tpmtotal'], self._format_run_note(best_mysql_tpmtotal)],
            ['mysql', 'summary phase', mysql_summary_runs[0]['phase'] or 'all', '优先使用start test; 若不存在则回退到全部轮次'],
            ['mysql', 'formal run count', len([run for run in mysql_runs if run['phase'] == 'start test']), 'phase=start test'],
            ['mysql', 'warmup run count', len([run for run in mysql_runs if run['phase'] == 'warmup']), 'phase=warmup'],
            ['mysql', 'total parsed run count', len(mysql_runs), 'mysql日志中的全部可解析轮次'],
            ['mysql', 'java heap mb', self.java_heap_mb, best_mysql_tpmc['java_options']],
            [
                'mysql', 'saved artifact dir count',
                len([item for item in self.mysql_artifacts.values() if item.get('saved_dir')]),
                str(self.directory / 'mysql_artifacts')
            ],
            ['postgresql', 'best measured tpmC', best_postgres_tpmc['measured_tpmc'], self._format_run_note(best_postgres_tpmc)],
            ['postgresql', 'best measured tpmTOTAL', best_postgres_tpmtotal['measured_tpmtotal'], self._format_run_note(best_postgres_tpmtotal)],
            ['postgresql', 'summary phase', postgres_summary_runs[0]['phase'] or 'all', '优先使用start test; 若不存在则回退到全部轮次'],
            ['postgresql', 'formal run count', len([run for run in postgres_runs if run['phase'] == 'start test']), 'phase=start test'],
            ['postgresql', 'warmup run count', len([run for run in postgres_runs if run['phase'] == 'warmup']), 'phase=warmup'],
            ['postgresql', 'total parsed run count', len(postgres_runs), 'postgresql日志中的全部可解析轮次'],
            ['postgresql', 'java heap mb', self.java_heap_mb, best_postgres_tpmc['java_options']],
            [
                'postgresql', 'saved artifact dir count',
                len([item for item in self.postgres_artifacts.values() if item.get('saved_dir')]),
                str(self.directory / 'postgresql_artifacts')
            ],
        ]
        for row in summary_rows:
            ws_summary.append(row)

        ws1 = wb.create_sheet(title='mysql')
        self._populate_db_sheet(ws1, mysql_runs)

        ws2 = wb.create_sheet(title='postgresql')
        self._populate_db_sheet(ws2, postgres_runs)

        ws3 = wb.create_sheet(title='build')
        self._populate_build_sheet(ws3, [self.mysql_build_info, self.postgres_build_info])

        for sheet in (ws_summary, ws1, ws2, ws3):
            sheet.freeze_panes = 'A2'
            for row in sheet.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(vertical='top', wrap_text=True)

        for sheet in (ws1, ws2):
            sheet.column_dimensions['A'].width = 14
            sheet.column_dimensions['B'].width = 12
            sheet.column_dimensions['C'].width = 24
            sheet.column_dimensions['D'].width = 48
            sheet.column_dimensions['E'].width = 12
            sheet.column_dimensions['F'].width = 12
            sheet.column_dimensions['G'].width = 12
            sheet.column_dimensions['H'].width = 12
            sheet.column_dimensions['I'].width = 18
            sheet.column_dimensions['J'].width = 12
            sheet.column_dimensions['K'].width = 16
            sheet.column_dimensions['L'].width = 20
            sheet.column_dimensions['M'].width = 14
            sheet.column_dimensions['N'].width = 14
            sheet.column_dimensions['O'].width = 18
            sheet.column_dimensions['P'].width = 14
            sheet.column_dimensions['Q'].width = 16
            sheet.column_dimensions['R'].width = 18
            sheet.column_dimensions['S'].width = 18
            sheet.column_dimensions['T'].width = 18
            sheet.column_dimensions['U'].width = 22
            sheet.column_dimensions['V'].width = 22
            sheet.column_dimensions['W'].width = 18
            sheet.column_dimensions['X'].width = 18
            sheet.column_dimensions['Y'].width = 18
            sheet.column_dimensions['Z'].width = 10
            sheet.column_dimensions['AA'].width = 28
            sheet.column_dimensions['AB'].width = 36
            sheet.column_dimensions['AC'].width = 36
            sheet.column_dimensions['AD'].width = 36
            sheet.column_dimensions['AE'].width = 36
            sheet.column_dimensions['AF'].width = 36
            sheet.column_dimensions['AG'].width = 28

        ws_summary.column_dimensions['A'].width = 14
        ws_summary.column_dimensions['B'].width = 24
        ws_summary.column_dimensions['C'].width = 18
        ws_summary.column_dimensions['D'].width = 42

        ws3.column_dimensions['A'].width = 14
        ws3.column_dimensions['B'].width = 24
        ws3.column_dimensions['C'].width = 48
        ws3.column_dimensions['D'].width = 12
        ws3.column_dimensions['E'].width = 12
        ws3.column_dimensions['F'].width = 12
        ws3.column_dimensions['G'].width = 26
        ws3.column_dimensions['H'].width = 30
        ws3.column_dimensions['I'].width = 12
        ws3.column_dimensions['J'].width = 18
        ws3.column_dimensions['K'].width = 20
        ws3.column_dimensions['L'].width = 12
        ws3.column_dimensions['M'].width = 12
        ws3.column_dimensions['N'].width = 12
        ws3.column_dimensions['O'].width = 16
        ws3.column_dimensions['P'].width = 18
        ws3.column_dimensions['Q'].width = 34
        ws3.column_dimensions['R'].width = 34

        wb.save(self.directory / 'benchmarksql.xlsx')


    def post_test(self):
        if hasattr(self, 'mysqld'):
            try:
                self.mysqld.Unit.Stop(b'replace')
            except Exception:
                pass
        if hasattr(self, 'postgresql'):
            try:
                self.postgresql.Unit.Stop(b'replace')
            except Exception:
                pass


    def run(self):
        print('开始进行benchmarksql测试')
        need_cleanup = False
        try:
            self.pre_test()
            need_cleanup = True
            self.run_test()
            try:
                self.result2summary()
            except Exception as e:
                logFile = self.directory / 'benchmarksql_summary_error.log'
                with open(logFile, 'w') as log:
                    log.write(str(e))
                raise SummaryError(logFile)
        finally:
            if need_cleanup or hasattr(self, 'mysqld') or hasattr(self, 'postgresql'):
                self.post_test()
        print('benchmarksql测试结束')

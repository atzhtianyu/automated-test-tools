from openpyxl.workbook import Workbook
from openpyxl.styles import Alignment
from pystemd.systemd1 import Unit
from pathlib import Path
from io import StringIO
import csv,time
import subprocess,shutil

from .errors import DefaultError,RunError,SummaryError


TEST_COMMANDS = (
    'PING_INLINE','PING_MBULK','SET','GET','INCR','LPUSH','RPUSH','LPOP','RPOP','SADD','HSET','SPOP','ZADD',
    'ZPOPMIN','LRANGE_100','LRANGE_300','LRANGE_500','LRANGE_600','MSET','XADD'
)
REDIS_BENCHMARK_HOST = '127.0.0.1'
REDIS_BENCHMARK_CLIENTS = 100
REDIS_BENCHMARK_REQUESTS = 100000


class redisBenchMark: # redis-benchmark 是 Redis 自带的基准测试工具
    def __init__(self, **kwargs):
        self.rpms = {'redis'}
        self.directory: Path = kwargs.get('saved_directory') / 'redis-benchmark'
        self.test_result:str = ''
        self.run_command = (
            f"redis-benchmark -h {REDIS_BENCHMARK_HOST} "
            f"-c {REDIS_BENCHMARK_CLIENTS} -n {REDIS_BENCHMARK_REQUESTS} --csv "
            f"-t {','.join(TEST_COMMANDS)}"
        )


    def pre_test(self):
        time.sleep(5)
        self.redis:Unit = Unit(b'redis.service',_autoload=True)
        try:
            self.redis.Unit.Start(b'replace')
        except:
            time.sleep(5)
            self.redis.load(force=True)
            self.redis.Unit.Start(b'replace')
        time.sleep(5)
        if self.redis.Unit.ActiveState != b'active':
            time.sleep(5)
            if self.redis.Unit.ActiveState != b'active':
                raise DefaultError("redis_benchmark测试出错.redis.service开启失败,退出测试.")

        try:
            subprocess.run(
                "command -v redis-benchmark",
                shell=True,check=True,
                stdout=subprocess.DEVNULL,stderr=subprocess.DEVNULL
            )
        except subprocess.CalledProcessError:
            raise DefaultError(f"redis_benchmark测试出错.找不到redis-benchmark命令.")


        if self.directory.exists():
            shutil.rmtree(self.directory)
        self.directory.mkdir(parents=True)


    def run_test(self):
        try:
            redis_bench_mark = subprocess.run(
                self.run_command,
                shell=True,check=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
            )
        except subprocess.CalledProcessError as e:
            raise RunError(e.returncode,e.stderr.decode())
        else:
            self.test_result = redis_bench_mark.stdout.decode('utf-8')
            with open(self.directory / 'redis-benchmark.csv', 'w', encoding='utf-8', newline='') as csv_file:
                csv_file.write(self.test_result)


    # 解析 redis-benchmark --csv 输出，转为含各延迟指标的结构化行列表
    def _parse_csv_rows(self):
        csv_reader = list(csv.reader(StringIO(self.test_result), delimiter=','))
        if len(csv_reader) <= 1:
            raise ValueError('redis-benchmark --csv 未返回有效测试结果')

        rows = []
        seen_test_names = set()
        for result in csv_reader[1:]:
            if not result or all(not item.strip() for item in result):
                continue
            if len(result) < 8:
                raise ValueError(f"redis-benchmark结果列数异常: {result}")
            test_name = result[0]
            if test_name in seen_test_names:
                raise ValueError(f"redis-benchmark结果中出现重复测试项: {test_name}")
            seen_test_names.add(test_name)

            rows.append({
                'test_name': test_name,
                'rps': float(result[1]),
                'avg_latency_ms': float(result[2]),
                'min_latency_ms': float(result[3]),
                'p50_latency_ms': float(result[4]),
                'p90_latency_ms': float(result[5]),
                'p99_latency_ms': float(result[6]),
                'max_latency_ms': float(result[7]),
            })

        if not rows:
            raise ValueError('redis-benchmark结果为空')
        return rows


    def result2symmary(self):
        parsed_rows = self._parse_csv_rows()

        wb = Workbook()
        ws_summary = wb.active
        ws_summary.title = 'summary'
        ws_summary.append(['类别', '指标', '值', '说明'])

        average_rps = sum(row['rps'] for row in parsed_rows) / len(parsed_rows)
        average_avg_latency = sum(row['avg_latency_ms'] for row in parsed_rows) / len(parsed_rows)
        average_p99_latency = sum(row['p99_latency_ms'] for row in parsed_rows) / len(parsed_rows)

        best_rps = max(parsed_rows, key=lambda row: row['rps'])
        worst_rps = min(parsed_rows, key=lambda row: row['rps'])
        lowest_avg_latency = min(parsed_rows, key=lambda row: row['avg_latency_ms'])
        highest_avg_latency = max(parsed_rows, key=lambda row: row['avg_latency_ms'])
        highest_p99_latency = max(parsed_rows, key=lambda row: row['p99_latency_ms'])
        highest_max_latency = max(parsed_rows, key=lambda row: row['max_latency_ms'])

        summary_rows = [
            ['测试参数', 'host', REDIS_BENCHMARK_HOST, 'redis-benchmark -h'],
            ['测试参数', 'parallel clients', REDIS_BENCHMARK_CLIENTS, 'redis-benchmark -c'],
            ['测试参数', 'total requests', REDIS_BENCHMARK_REQUESTS, 'redis-benchmark -n'],
            ['测试参数', 'tested commands', len(parsed_rows), '本次实际落表的命令数量'],
            ['测试参数', 'test command', self.run_command, '原始执行命令'],
            ['汇总指标', '平均 requests/sec', round(average_rps, 2), '越大越好'],
            ['汇总指标', '平均 avg latency(ms)', round(average_avg_latency, 4), '越小越好'],
            ['汇总指标', '平均 p99 latency(ms)', round(average_p99_latency, 4), '越小越好'],
            ['汇总指标', '最高吞吐命令', best_rps['test_name'], f"{best_rps['rps']:.2f} requests/sec"],
            ['汇总指标', '最低吞吐命令', worst_rps['test_name'], f"{worst_rps['rps']:.2f} requests/sec"],
            ['汇总指标', '最低平均延迟命令', lowest_avg_latency['test_name'], f"{lowest_avg_latency['avg_latency_ms']:.4f} ms"],
            ['汇总指标', '最高平均延迟命令', highest_avg_latency['test_name'], f"{highest_avg_latency['avg_latency_ms']:.4f} ms"],
            ['汇总指标', '最高 p99 延迟命令', highest_p99_latency['test_name'], f"{highest_p99_latency['p99_latency_ms']:.4f} ms"],
            ['汇总指标', '最高最大延迟命令', highest_max_latency['test_name'], f"{highest_max_latency['max_latency_ms']:.4f} ms"],
        ]
        for row in summary_rows:
            ws_summary.append(row)

        ws_ranking = wb.create_sheet(title='ranking')
        ws_ranking.append(['排名', '测试项目名称', 'rps(每秒请求数)', '平均延迟(ms)', '99% 延迟(ms)', '最大延迟(ms)'])
        for index, row in enumerate(sorted(parsed_rows, key=lambda item: item['rps'], reverse=True), start=1):
            ws_ranking.append([
                index,
                row['test_name'],
                row['rps'],
                row['avg_latency_ms'],
                row['p99_latency_ms'],
                row['max_latency_ms'],
            ])

        ws = wb.create_sheet(title='redisBenchmark')
        ws.append(['测试项目名称','rps(每秒请求数)','平均延迟(ms)','最小延迟(ms)','50% 延迟(ms)[中位数]','90% 延迟(ms)','99% 延迟(ms)','最大延迟(ms)'])
        for row in parsed_rows:
            ws.append([
                row['test_name'],
                row['rps'],
                row['avg_latency_ms'],
                row['min_latency_ms'],
                row['p50_latency_ms'],
                row['p90_latency_ms'],
                row['p99_latency_ms'],
                row['max_latency_ms'],
            ])

        for sheet in (ws_summary, ws_ranking, ws):
            sheet.freeze_panes = 'A2'
            for row in sheet.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(vertical='top', wrap_text=True)

        ws_summary.column_dimensions['A'].width = 12
        ws_summary.column_dimensions['B'].width = 24
        ws_summary.column_dimensions['C'].width = 28
        ws_summary.column_dimensions['D'].width = 32

        ws_ranking.column_dimensions['A'].width = 8
        ws_ranking.column_dimensions['B'].width = 18
        for column in ('C', 'D', 'E', 'F'):
            ws_ranking.column_dimensions[column].width = 18

        ws.column_dimensions['A'].width = 18
        for column in ('B', 'C', 'D', 'E', 'F', 'G', 'H'):
            ws.column_dimensions[column].width = 18

        wb.save(self.directory / 'redis-benchmark.xlsx')


    def post_test(self):
        self.redis.Unit.Stop(b'replace')
        time.sleep(5)
        subprocess.run(
            "dnf remove -y redis",shell=True,stdout=subprocess.DEVNULL,stderr=subprocess.DEVNULL
        )


    def run(self):
        print('开始进行redis_benchmark测试')
        self.pre_test()
        self.run_test()
        try:
            self.result2symmary()
        except Exception as e:
            logFile = self.directory / 'redisBenchmark_summary_error.log'
            with open(logFile,'w') as log:
                log.write(str(e))
            raise SummaryError(logFile)
        self.post_test()
        print('redis_benchmark测试结束')

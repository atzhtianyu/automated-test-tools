from openpyxl.workbook import Workbook
from openpyxl.styles import Alignment
from pystemd.systemd1 import Unit
from io import BytesIO
from pathlib import Path
import re
import shlex
import pexpect
import time
import requests,tarfile
import pymysql
import subprocess,shutil
from tqdm import trange,tqdm

from .errors import DefaultError,SummaryError


class TPC_H:
    def __init__(self, **kwargs):
        self.rpms = {'sysbench','mysql-server','mysql'}
        self.directory: Path = kwargs.get('saved_directory') / 'TPC-H'
        self.path = Path('/root/osmts_tmp/TPC-H')
        self.dbgen = self.path / 'dbgen'
        self.saveSQL = self.dbgen / 'saveSQL'
        self.sql_results:list = []


    # 将 qgen 生成的原生 TPC-H SQL 自动修改为 MySQL 兼容版本
    def _patch_mysql_query_files(self):
        # 文档要求将qgen生成的原生TPC-H SQL手动调整为MySQL兼容版本。
        # 这里把这些固定改动自动化，避免查询阶段直接报语法错误。
        for sql_file in sorted(self.saveSQL.glob('*.sql')):
            content = sql_file.read_text(encoding='utf-8')
            content = content.replace('limit -1;', '')

            if sql_file.name == '1.sql':
                content = content.replace('day (3)', 'day')

            if sql_file.name in {'2.sql', '3.sql', '10.sql', '18.sql', '21.sql'}:
                lines = content.splitlines()
                non_empty_indexes = [index for index, line in enumerate(lines) if line.strip()]
                if len(non_empty_indexes) >= 2:
                    penultimate_index = non_empty_indexes[-2]
                    lines[penultimate_index] = re.sub(r';\s*$', '', lines[penultimate_index])
                    content = '\n'.join(lines)
                    if not content.endswith('\n'):
                        content += '\n'

            sql_file.write_text(content, encoding='utf-8')


    # 将 "X hour Y min Z sec" 格式的耗时文本转换为秒数
    def _duration_to_seconds(self, duration_text: str) -> float:
        total_seconds = 0.0
        matched = False

        hour_match = re.search(r'(\d+)\s+hour', duration_text)
        minute_match = re.search(r'(\d+)\s+min', duration_text)
        second_match = re.search(r'([\d.]+)\s+sec', duration_text)

        if hour_match:
            matched = True
            total_seconds += int(hour_match.group(1)) * 3600
        if minute_match:
            matched = True
            total_seconds += int(minute_match.group(1)) * 60
        if second_match:
            matched = True
            total_seconds += float(second_match.group(1))

        if not matched:
            raise ValueError(f'无法解析TPC-H查询耗时: {duration_text}')

        return round(total_seconds, 2)


    # 从 mysql 输出中解析单条 SQL 的返回行数和执行耗时
    def _parse_sql_result(self, sql_index: int, sql_result: str):
        rows_match = re.search(r'(\d+)\s+rows?\s+in\s+set\s+\(([^)]+)\)', sql_result)
        empty_match = re.search(r'Empty\s+set\s+\(([^)]+)\)', sql_result)

        if rows_match:
            rows_count = int(rows_match.group(1))
            duration_text = rows_match.group(2).strip()
        elif empty_match:
            rows_count = 0
            duration_text = empty_match.group(1).strip()
        else:
            error_match = re.search(r'(ERROR\s+\d+\s+\([^)]+\):.+)', sql_result)
            if error_match:
                raise ValueError(f'第{sql_index}条TPC-H SQL执行失败: {error_match.group(1)}')
            snippet = '\n'.join(line for line in sql_result.splitlines()[-12:] if line.strip())
            raise ValueError(f'无法在TPC-H查询结果中找到第{sql_index}条SQL的耗时信息.输出片段:\n{snippet}')

        return {
            'sql_file': f'{sql_index}.sql',
            'rows_count': rows_count,
            'elapsed_text': duration_text,
            'elapsed_seconds': self._duration_to_seconds(duration_text),
        }


    # 检查 mysql 客户端命令是否可用，不可用则抛出明确的错误信息
    def _ensure_mysql_client(self):
        mysql_path = shutil.which('mysql')
        if mysql_path:
            return mysql_path
        raise DefaultError(
            "TPC-H测试出错.未找到mysql客户端命令.当前脚本需要mysql CLI来执行SOURCE和\\.命令."
            "请确认mysql-server已正确安装并提供mysql命令; 如果当前系统将客户端单独打包,请额外安装对应的mysql客户端包后重试."
        )


    def pre_test(self):
        if self.directory.exists():
            shutil.rmtree(self.directory)
        self.directory.mkdir(parents=True)
        if self.path.exists():
            shutil.rmtree(self.path)

        time.sleep(5)
        self.mysqld:Unit = Unit('mysqld.service',_autoload=True)
        try:
            self.mysqld.Unit.Start(b'replace')
        except Exception:
            time.sleep(5)
            self.mysqld.load(force=True)
            self.mysqld.Unit.Start(b'replace')
        time.sleep(5)
        if self.mysqld.Unit.ActiveState != b'active':
            time.sleep(5)
            if self.mysqld.Unit.ActiveState != b'active':
                raise DefaultError("sysbench测试出错.开启mysqld.service失败,退出测试.")

        try:
            self.conn = pymysql.connect(
                host='localhost',
                port=3306,
                user='root',
                passwd='',
            )
        except Exception as e:
            self.conn = pymysql.connect(
                host='localhost',
                port=3306,
                user='root',
                passwd='123456',
            )
        cursor = self.conn.cursor()
        cursor.execute("ALTER USER 'root'@'localhost' IDENTIFIED BY '123456';")

        # 获取TPC-H
        response = requests.get(
            url="https://gitee.com/April_Zhao/osmts/releases/download/v1.0/TPC-H.tar.xz",
            headers={
                'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64; rv:137.0) Gecko/20100101 Firefox/137.0',
                'referer': 'https://gitee.com/April_Zhao/osmts',
            }
        )
        response.raise_for_status()
        with tarfile.open(fileobj=BytesIO(response.content), mode="r:xz") as tar:
            tar.extractall(Path('/root/osmts_tmp/'))

        # build dbgen
        # 这个过程会有交互
        try:
            subprocess.run(
                "make -j 4 && ./dbgen -s 1",
                cwd=self.path / "dbgen",
                shell=True,check=True,
                stdout=subprocess.DEVNULL,
                stderr=subprocess.PIPE,
            )
        except subprocess.CalledProcessError as e:
            raise DefaultError(f"TPC-H测试出错.构建或运行dbgen失败,报错信息:{e.stdout.decode('utf-8')}")

        try:
            subprocess.run(
                "cp -f qgen dists.dss queries/ && cd queries &&"
                "for i in {1..22};do ./qgen -d ${i} > ../saveSQL/${i}.sql;done",
                cwd=self.path / "dbgen",
                shell=True,check=True,
                stdout=subprocess.DEVNULL,
                stderr=subprocess.PIPE,
            )
        except subprocess.CalledProcessError as e:
            raise DefaultError(f"TPC-H测试出错.生成saveSQL失败,报错信息:{e.stdout.decode('utf-8')}")

        self._patch_mysql_query_files()



        cursor.execute("DROP DATABASE IF EXISTS tpch;")
        cursor.execute("CREATE DATABASE IF NOT EXISTS tpch;")
        cursor.execute("USE tpch;")
        cursor.close()

        # SOURCE是MySQL客户端特有的工具(pymysql无法执行SOURCE)
        # cursor.execute(f"SOURCE {self.path}/dss.ddl;")
        # cursor.execute(f"SOURCE {self.path}/dss.ri;")
        self.mysql_client = self._ensure_mysql_client()


        mysql = pexpect.spawn(
            command="/bin/bash",
            args=["-c", f"{shlex.quote(self.mysql_client)} -uroot -p123456"],
            encoding='utf-8',
            logfile=open(self.directory / 'osmts_tpch_loadData.log', 'w'),
        )
        mysql.expect_exact("mysql>", timeout=60)
        mysql.sendline(f"USE tpch;")

        mysql.expect_exact("mysql>", timeout=60)
        # mysql服务器端和客户端在一次传送数据包的过程当中最大允许的数据包大小
        mysql.sendline("SET GLOBAL max_allowed_packet = 1024*1024*1024;")

        # 用于缓存索引和数据的内存大小
        mysql.expect_exact("mysql>", timeout=60)
        mysql.sendline("SET GLOBAL innodb_buffer_pool_size = 4*1024*1024*1024;")

        mysql.expect_exact("mysql>", timeout=60)
        mysql.sendline(f"SOURCE {self.dbgen}/dss.ddl;")
        mysql.expect_exact("mysql>", timeout=600)
        mysql.sendline(f"SOURCE {self.dbgen}/dss.ri;")

        mysql.expect_exact("mysql>", timeout=600)
        mysql.sendline("SET FOREIGN_KEY_CHECKS=0;")
        mysql.expect_exact("mysql>", timeout=60)
        mysql.sendline("SET GLOBAL local_infile=1;")
        mysql.expect_exact("mysql>", timeout=60)

        for table in tqdm(('customer','lineitem','nation','orders','partsupp','part','region','supplier'),desc="load data进度"):
            mysql.sendline(
                f"LOAD DATA LOCAL INFILE '{self.dbgen}/{table}.tbl' INTO TABLE {table} FIELDS TERMINATED BY '|' LINES TERMINATED BY '|\n';"
            )
            mysql.expect_exact("mysql>", timeout=3600)

        mysql.sendline("SET FOREIGN_KEY_CHECKS=1;")
        mysql.expect_exact("mysql>")
        mysql.terminate(force=True)



    def run_test(self):
        mysql = pexpect.spawn(
            command="/bin/bash",
            args=["-c", f"{shlex.quote(self.mysql_client)} -uroot -p123456"],
            encoding='utf-8',
            logfile=open(self.directory / 'osmts_tpch_query.log', 'w'),
        )
        mysql.expect_exact("mysql>", timeout=60)
        mysql.sendline(f"USE tpch;")
        mysql.expect_exact("mysql>", timeout=60)

        for i in trange(1,23,desc="SQL查询进度"):
            mysql.sendline(f"\\. {self.saveSQL}/{i}.sql")
            mysql.expect_exact("mysql>", timeout=36000)
            self.sql_results.append(mysql.before)
        time.sleep(5)
        mysql.terminate(force=True)


    def result2summary(self):
        parsed_results = [
            self._parse_sql_result(index, sql_result)
            for index, sql_result in enumerate(self.sql_results, start=1)
        ]

        wb = Workbook()
        ws_summary = wb.active
        ws_summary.title = 'summary'
        ws_summary.append(['metric', 'value', 'note'])

        total_elapsed = round(sum(item['elapsed_seconds'] for item in parsed_results), 2)
        average_elapsed = round(total_elapsed / len(parsed_results), 2)
        slowest_query = max(parsed_results, key=lambda item: item['elapsed_seconds'])
        fastest_query = min(parsed_results, key=lambda item: item['elapsed_seconds'])
        largest_result = max(parsed_results, key=lambda item: item['rows_count'])

        ws_summary.append(['query count', len(parsed_results), 'TPC-H标准查询数量'])
        ws_summary.append(['total elapsed time(s)', total_elapsed, '22条SQL累计耗时'])
        ws_summary.append(['average elapsed time(s)', average_elapsed, '平均每条SQL耗时'])
        ws_summary.append(['slowest query', slowest_query['sql_file'], f"{slowest_query['elapsed_seconds']}s"])
        ws_summary.append(['fastest query', fastest_query['sql_file'], f"{fastest_query['elapsed_seconds']}s"])
        ws_summary.append(['largest result set query', largest_result['sql_file'], f"{largest_result['rows_count']} rows"])

        ws = wb.create_sheet(title='TPC-H')
        ws.title = 'TPC-H'
        ws.append(['SQL文件', '返回行数', '查询所耗时间', '查询所耗时间(s)'])
        for item in parsed_results:
            ws.append([
                item['sql_file'],
                item['rows_count'],
                item['elapsed_text'],
                item['elapsed_seconds'],
            ])

        ws_summary.freeze_panes = 'A2'
        ws.freeze_panes = 'A2'
        for sheet in (ws_summary, ws):
            for row in sheet.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(vertical='top', wrap_text=True)

        ws_summary.column_dimensions['A'].width = 24
        ws_summary.column_dimensions['B'].width = 18
        ws_summary.column_dimensions['C'].width = 24

        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 16

        wb.save(self.directory / 'tpch.xlsx')


    def post_test(self):
        if hasattr(self, 'mysqld'):
            try:
                self.mysqld.Unit.Stop(b'replace')
            except Exception:
                pass


    def run(self):
        print('开始进行tpch测试')
        self.pre_test()
        self.run_test()
        try:
            self.result2summary()
        except Exception as e:
            logFile = self.directory / 'tpch_summary_error.log'
            with open(logFile, 'w') as log:
                log.write(str(e))
            raise SummaryError(logFile)
        self.post_test()
        print('tpch测试结束')

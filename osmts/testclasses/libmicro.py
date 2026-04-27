from pathlib import Path
import subprocess,shutil,re
from openpyxl import Workbook

from .errors import GitCloneError,CompileError,RunError


class Libmicro:
    def __init__(self, **kwargs):
        self.rpms = set()
        self.believe_tmp: bool = kwargs.get('believe_tmp')
        self.path = Path('/root/osmts_tmp/libmicro')
        self.directory: Path = kwargs.get('saved_directory') / 'libmicro'
        self.compiler: str = kwargs.get('compiler')
        self.test_result = ''


    def pre_test(self):
        if self.path.exists() and self.believe_tmp:
            pass
        else:
            shutil.rmtree(self.path, ignore_errors=True)
            # 获取源码
            try:
                subprocess.run(
                    args="git clone https://gitee.com/zhtianyu/libmicro.git",
                    cwd="/root/osmts_tmp",
                    shell=True,check=True,
                    stdout=subprocess.DEVNULL,
                    stderr=subprocess.PIPE
                )
            except subprocess.CalledProcessError as e:
                raise GitCloneError(e.returncode,'https://gitee.com/zhtianyu/libmicro.git',e.stderr.decode())


        # 开始编译
        try:
            if self.compiler == "gcc":
                subprocess.run(
                    args="make",
                    cwd=self.path,
                    shell=True,check=True,
                    stdout=subprocess.DEVNULL,
                    stderr=subprocess.PIPE
                )
            elif self.compiler == "clang":
                subprocess.run(
                    args='make CC=clang CFLAGS="-Wno-error=implicit-function-declaration"',
                    cwd=self.path,
                    shell=True,check=True,
                    stdout=subprocess.DEVNULL,
                    stderr=subprocess.PIPE
                )
        except subprocess.CalledProcessError as e:
            raise CompileError(e.returncode,self.compiler,e.stderr.decode())


    def run_test(self):
        try:
            bench = subprocess.run(
                args="./bench",
                cwd=self.path,
                shell=True,check=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE
            )
        except subprocess.CalledProcessError as e:
            raise RunError(e.returncode,e.stderr.decode())
        else:
            self.test_result = bench.stdout.decode('utf-8')
        if not self.directory.exists():
            self.directory.mkdir(exist_ok=True,parents=True)
        with open(self.directory / 'libmicro.log','w') as file:
            file.write(self.test_result)


    def result2summary(self):
        metadata = {}
        benches = []
        lines = self.test_result.splitlines()
        i = 0

        while i < len(lines):
            line = lines[i]
            if line.startswith('!') and ':' in line:
                key, value = line[1:].split(':', 1)
                metadata[key.strip()] = value.strip()
                i += 1
                continue

            if line.startswith('# bin/'):
                command = line[2:].strip()
                bench = {
                    'command': command,
                    'name': '',
                    'prc': '',
                    'thr': '',
                    'usecs_per_call': '',
                    'samples': '',
                    'errors': '',
                    'cnt_samp': '',
                    'raw_min': '',
                    'raw_max': '',
                    'raw_mean': '',
                    'raw_median': '',
                    'raw_stddev': '',
                    'raw_confidence_99': '',
                    'elapsed_time': '',
                    'outliers': '',
                    'percentile_95': '',
                }

                i += 1
                while i < len(lines):
                    current = lines[i].strip()
                    if not current:
                        i += 1
                        continue
                    if current.startswith('# bin/'):
                        break

                    row_match = re.match(
                        r'^(\S+)\s+(\d+)\s+(\d+)\s+([\d.]+)\s+(\d+)\s+(\d+)\s+(\d+)$',
                        current
                    )
                    if row_match:
                        bench['name'] = row_match.group(1)
                        bench['prc'] = int(row_match.group(2))
                        bench['thr'] = int(row_match.group(3))
                        bench['usecs_per_call'] = float(row_match.group(4))
                        bench['samples'] = int(row_match.group(5))
                        bench['errors'] = int(row_match.group(6))
                        bench['cnt_samp'] = int(row_match.group(7))
                        i += 1
                        continue

                    stat_match = re.match(
                        r'^#\s+(min|max|mean|median|stddev|99% confidence level)\s+([\d.]+)\s+([\d.]+)$',
                        current
                    )
                    if stat_match:
                        stat_name = stat_match.group(1)
                        raw_value = float(stat_match.group(2))
                        if stat_name == 'min':
                            bench['raw_min'] = raw_value
                        elif stat_name == 'max':
                            bench['raw_max'] = raw_value
                        elif stat_name == 'mean':
                            bench['raw_mean'] = raw_value
                        elif stat_name == 'median':
                            bench['raw_median'] = raw_value
                        elif stat_name == 'stddev':
                            bench['raw_stddev'] = raw_value
                        elif stat_name == '99% confidence level':
                            bench['raw_confidence_99'] = raw_value
                        i += 1
                        continue

                    elapsed_match = re.match(r'^#\s+elasped time\s+([\d.]+)$', current)
                    if elapsed_match:
                        bench['elapsed_time'] = float(elapsed_match.group(1))
                        i += 1
                        continue

                    outliers_match = re.match(r'^#\s+number of outliers\s+(\d+)$', current)
                    if outliers_match:
                        bench['outliers'] = int(outliers_match.group(1))
                        i += 1
                        continue

                    percentile_match = re.match(r'^#\s+95th %ile\s+([\d.]+)$', current)
                    if percentile_match:
                        bench['percentile_95'] = float(percentile_match.group(1))
                        i += 1
                        continue

                    i += 1

                benches.append(bench)
                continue

            i += 1

        wb = Workbook()
        ws = wb.active
        ws.title = 'summary'
        ws.append(['项目', '值'])

        summary_items = [
            ('Libmicro版本', metadata.get('Libmicro_#', '')),
            ('编译器', metadata.get('Compiler', self.compiler)),
            ('编译器版本', metadata.get('Compiler Ver.', '')),
            ('机器名', metadata.get('Machine_name', '')),
            ('OS版本', metadata.get('OS_release', '')),
            ('CPU架构', metadata.get('Processor', '')),
            ('CPU数量', metadata.get('#CPUs', '')),
            ('运行用户', metadata.get('Run_by', '')),
            ('测试时间', metadata.get('Date', '')),
            ('测试选项', metadata.get('Options', '')),
            ('样例总数', len(benches)),
            ('错误样例数', sum(1 for bench in benches if bench['errors'])),
            ('总errors', sum(bench['errors'] for bench in benches if bench['errors'] != '')),
            ('平均usecs/call', round(sum(bench['usecs_per_call'] for bench in benches if bench['usecs_per_call'] != '') / len([bench for bench in benches if bench['usecs_per_call'] != '']), 5) if benches else ''),
            ('日志文件', 'libmicro.log'),
        ]
        for item, value in summary_items:
            ws.append([item, value])

        wb.create_sheet(title='benchmarks')
        ws = wb['benchmarks']
        ws.append([
            '测试项', '进程数', '线程数', 'usecs/call', 'samples', 'errors', 'cnt/samp',
            'raw min', 'raw max', 'raw mean', 'raw median', 'raw stddev',
            '99% confidence', 'elapsed time', 'outliers', '95th %ile', '命令'
        ])
        for bench in benches:
            ws.append([
                bench['name'],
                bench['prc'],
                bench['thr'],
                bench['usecs_per_call'],
                bench['samples'],
                bench['errors'],
                bench['cnt_samp'],
                bench['raw_min'],
                bench['raw_max'],
                bench['raw_mean'],
                bench['raw_median'],
                bench['raw_stddev'],
                bench['raw_confidence_99'],
                bench['elapsed_time'],
                bench['outliers'],
                bench['percentile_95'],
                bench['command'],
            ])

        wb.create_sheet(title='top slowest')
        ws = wb['top slowest']
        ws.append(['排名', '测试项', 'usecs/call', '95th %ile', 'errors'])
        sorted_benches = sorted(
            [bench for bench in benches if bench['usecs_per_call'] != ''],
            key=lambda item: item['usecs_per_call'],
            reverse=True
        )
        for index, bench in enumerate(sorted_benches[:50], start=1):
            ws.append([
                index,
                bench['name'],
                bench['usecs_per_call'],
                bench['percentile_95'],
                bench['errors'],
            ])

        wb.save(self.directory / 'libmicro.xlsx')


    def run(self):
        print("开始进行libmicro测试")
        self.pre_test()
        self.run_test()
        self.result2summary()
        print("libmicro测试结束")

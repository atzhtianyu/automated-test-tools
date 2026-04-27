import re
import shutil
import subprocess
from pathlib import Path
from openpyxl import Workbook

from .errors import GitCloneError, RunError


class DejaGnu:
    def __init__(self, **kwargs):
        self.rpms = {'gcc-g++', 'gcc-gfortran', 'dejagnu'}
        self.path = Path('/root/osmts_tmp/dejagnu')
        self.directory: Path = kwargs.get('saved_directory') / 'dejagnu'
        self.testsuite = Path('/root/osmts_tmp/dejagnu/gcc/gcc/testsuite/')

    def pre_test(self):
        if not self.directory.exists():
            self.directory.mkdir(parents=True, exist_ok=True)
        if self.path.exists():
            shutil.rmtree(self.path)
        self.path.mkdir(parents=True)
        try:
            subprocess.run(
                'git clone https://gitee.com/openeuler/gcc.git',
                cwd=self.path,
                shell=True, check=True,
                stdout=subprocess.DEVNULL,
                stderr=subprocess.PIPE,
            )
        except subprocess.CalledProcessError as e:
            raise GitCloneError(
                e.returncode, 'https://gitee.com/openeuler/gcc.git',
                e.stderr.decode('utf-8') if e.stderr else '',
            )

    # 解析 dejagnu .sum 文件，将测试结果按 PASS/FAIL/ERROR 等状态分类归纳
    def _parse_sum(self, sum_path):
        result = {'PASS': [], 'FAIL': [], 'UNSUPPORTED': [],
                  'XFAIL': [], 'XPASS': [], 'WARNING': [], 'ERROR': [], 'other': []}
        summary = {}
        with open(sum_path, errors='ignore') as f:
            for line in f:
                line = line.rstrip()
                m = re.match(r'^(PASS|FAIL|UNSUPPORTED|XFAIL|XPASS|WARNING|ERROR|KFAIL):\s+(.*)$', line)
                if m:
                    status = m.group(1)
                    result[status].append(m.group(2))
                    continue
                sm = re.match(r'^# of (\S+(?:\s+\S+)*)\s+(\d+)$', line)
                if sm:
                    summary[sm.group(1)] = int(sm.group(2))
        return result, summary

    # 将 dejagnu 测试结果按状态优先级写入 Excel 工作表，并附上统计汇总
    def _write_excel(self, results, summary, tool):
        wb = Workbook()
        ws = wb.active
        ws.title = tool

        ws.append(['状态', '测试用例', '详情'])
        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 80

        row = 2
        for status in ('FAIL', 'XPASS', 'ERROR', 'UNSUPPORTED', 'UNRESOLVED', 'WARNING', 'XFAIL', 'PASS'):
            for case in results.get(status, []):
                case = case.strip()
                m = re.match(r'^(\S+)\s+(.*)$', case)
                test_case = m.group(1) if m else case
                detail = m.group(2).strip() if m else ''
                ws.cell(row=row, column=1, value=status)
                ws.cell(row=row, column=2, value=test_case)
                ws.cell(row=row, column=3, value=detail)
                row += 1

        if summary:
            row += 1
            ws.cell(row=row, column=1, value='汇总统计')
            for key, val in summary.items():
                row += 1
                ws.cell(row=row, column=1, value=key)
                ws.cell(row=row, column=2, value=val)

        wb.save(self.directory / f'{tool}.xlsx')
        print(f"[dejagnu] 已生成 {tool} Excel 报表: {self.directory / f'{tool}.xlsx'}")

    def run_test(self):
        for tool, logname in [('gcc', 'gcc'), ('g++', 'g++'), ('gfortran', 'gfortran')]:
            print(f"[dejagnu] 开始运行 {tool} 测试")
            result = subprocess.run(
                f'runtest --tool {tool}',
                cwd=self.testsuite,
                shell=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
            )

            stderr_output = result.stderr.decode('utf-8') if result.stderr else ''
            if result.returncode != 0 and "Couldn't find the global config file." not in stderr_output:
                raise RunError(
                    result.returncode,
                    f"dejagnu测试出错.runtest --tool {tool} 命令运行失败,报错信息:\n{stderr_output}",
                )

            remaining = '\n'.join([
                line for line in stderr_output.splitlines()
                if "Couldn't find the global config file." not in line
            ])
            if remaining.strip():
                print(f"[{tool}] stderr 警告信息:\n{remaining}")

            log_file = self.testsuite / f'{logname}.log'
            sum_file = self.testsuite / f'{logname}.sum'
            if log_file.exists() and sum_file.exists():
                shutil.copy(log_file, self.directory)
                shutil.copy(sum_file, self.directory)
            else:
                print(f"[警告] 未生成 {logname}.log 或 {logname}.sum 文件，可能测试未运行")
                continue

            test_results, summary = self._parse_sum(sum_file)
            self._write_excel(test_results, summary, tool)

            total_pass = len(test_results.get('PASS', []))
            total_fail = len(test_results.get('FAIL', []))
            total_unsupported = len(test_results.get('UNSUPPORTED', []))
            print(f"[dejagnu] {tool} 测试完成: PASS={total_pass} FAIL={total_fail} "
                  f"UNSUPPORTED={total_unsupported}")

    def run(self):
        print('开始进行dejagnu测试')
        self.pre_test()
        self.run_test()
        print('dejagnu测试结束')
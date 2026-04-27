import re
import subprocess
from pathlib import Path
from typing import Optional
from openpyxl import Workbook
from .errors import DefaultError



class OpenSCAP:
    def __init__(self, **kwargs):
        self.rpms = {'openscap','scap-security-guide'}
        self.scap_content = Path("/usr/share/xml/scap/ssg/content/")
        self.directory: Path = kwargs.get('saved_directory') / 'OpenSCAP'
        self.profile_id = "xccdf_org.ssgproject.content_profile_standard"
        self.datastream_path: Optional[Path] = None
        self.test_result:str = ''
        self.reinforce_result:str = ''


    # 根据系统版本号自动定位对应的 SCAP 数据流 XML 文件路径
    def _resolve_datastream_path(self):
        release_text = ''
        try:
            with open('/etc/openEuler-release','r', encoding='utf-8') as file:
                release_text = file.read().strip()
        except FileNotFoundError:
            release_text = 'openEuler 24.03'

        match = re.search(r'(\d{2})\.(\d{2})', release_text)
        if match:
            candidate = self.scap_content / f"ssg-openeuler{match.group(1)}{match.group(2)}-ds.xml"
            if candidate.exists():
                return candidate

        fallback = self.scap_content / "ssg-openeuler2403-ds.xml"
        if fallback.exists():
            return fallback

        candidates = sorted(self.scap_content.glob('ssg-openeuler*03-ds.xml'))
        if candidates:
            return candidates[-1]

        raise DefaultError("OpenSCAP测试出错.未找到可用的SCAP数据流文件.")


    # 从 oscap 文本输出中提取 (title, rule, result) 三元组列表
    def _extract_result_rows(self, content: str):
        titles = re.findall(r"Title {3}(.+)",content)
        rules = re.findall(r"Rule {3}(.+)",content)
        results = re.findall(r"Result {3}(.+)",content)
        row_count = min(len(titles), len(rules), len(results))
        return [(titles[i], rules[i], results[i]) for i in range(row_count)]


    # 将 stdout、stderr 和 returncode 格式化为结构化的日志文本
    def _format_process_output(self, stdout: str, stderr: str, returncode: int):
        output_parts = [f"[returncode]\n{returncode}"]
        if stdout.strip():
            output_parts.append(stdout.rstrip())
        if stderr.strip():
            output_parts.append(f"[stderr]\n{stderr.rstrip()}")
        return '\n\n'.join(output_parts) + '\n'


    # 执行 oscap 命令并将格式化后的输出写入日志文件
    def _run_oscap(self, command: str, error_prefix: str, log_path: Path):
        oscap = subprocess.run(
            command,
            shell=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
        )
        stdout = oscap.stdout.decode('utf-8', errors='replace')
        stderr = oscap.stderr.decode('utf-8', errors='replace')
        combined_output = self._format_process_output(stdout, stderr, oscap.returncode)

        with open(log_path, 'w', encoding='utf-8') as file:
            file.write(combined_output)

        # oscap xccdf eval:
        # 0 = all rules pass
        # 2 = evaluation completed, but at least one rule failed/unknown
        if oscap.returncode not in (0, 2):
            detail = stderr.strip() or stdout.strip() or f"oscap返回码:{oscap.returncode}"
            raise DefaultError(f"{error_prefix}{detail}")

        return stdout if stdout.strip() else combined_output


    def pre_test(self):
        if not self.directory.exists():
            self.directory.mkdir(parents=True)
        self.datastream_path = self._resolve_datastream_path()


    def run_test(self):
        self.test_result = self._run_oscap(
            f"oscap xccdf eval "
            f"--profile {self.profile_id} "
            f"--results {self.directory}/scan_results.xml "
            f"--report {self.directory}/scan_report.html "
            f"{self.datastream_path}",
            "OpenSCAP测试出错.基础合规检查失败,报错信息:",
            self.directory / 'openscap.log',
        )

        self.reinforce_result = self._run_oscap(
            f"oscap xccdf eval --remediate "
            f"--profile {self.profile_id} "
            f"--results {self.directory}/scan_results_reinforce.xml "
            f"--report {self.directory}/scan_report_reinforce.html "
            f"{self.datastream_path}",
            "OpenSCAP测试出错.加固复测失败,报错信息:",
            self.directory / 'openscap_reinforce.log',
        )


    def result2summary(self):
        wb = Workbook()
        baseline_rows = self._extract_result_rows(self.test_result)
        reinforce_rows = self._extract_result_rows(self.reinforce_result)

        ws_summary = wb.active
        ws_summary.title = 'summary'
        ws_summary.append(['metric', 'value'])
        ws_summary.append(['datastream', str(self.datastream_path)])
        ws_summary.append(['profile', self.profile_id])
        ws_summary.append(['baseline rows', len(baseline_rows)])
        ws_summary.append(['reinforce rows', len(reinforce_rows)])

        ws = wb.create_sheet(title='baseline')
        ws.append(['Title','Rule','Result'])
        if baseline_rows:
            for row in baseline_rows:
                ws.append(list(row))
        else:
            ws.append(['(无可解析结果)', '', ''])

        ws_reinforce = wb.create_sheet(title='reinforce')
        ws_reinforce.append(['Title','Rule','Result'])
        if reinforce_rows:
            for row in reinforce_rows:
                ws_reinforce.append(list(row))
        else:
            ws_reinforce.append(['(无可解析结果)', '', ''])

        wb.save(self.directory / 'openscap.xlsx')


    def run(self):
        print('开始进行OpenSCAP测试')
        self.pre_test()
        self.run_test()
        self.result2summary()
        print('OpenSCAP测试结束')

from pathlib import Path
import os,signal,sys,subprocess,re,time
from collections import Counter, defaultdict, deque
from openpyxl import Workbook

from .errors import DefaultError,GitCloneError


class Trinity:
    def __init__(self,**kwargs ):
        self.rpms = set()
        self.directory:Path = kwargs.get('saved_directory') / 'trinity'
        self.compiler:str = kwargs.get('compiler')
        self.test_result:str = ''
        self.user_home = Path('/home/trinity_test')
        self.trinity_path = self.user_home / f'trinity_{self.compiler}'
        self.log_path = self.directory / 'trinity.log'
        self.report_path = self.directory / 'trinity_report.xlsx'


    # 获取 trinity_test 用户的所有进程 PID 列表
    def _get_user_pids(self):
        result = subprocess.run(
            ['ps', '-u', 'trinity_test', '-o', 'pid='],
            stdout=subprocess.PIPE,
            stderr=subprocess.DEVNULL,
            text=True,
        )
        if result.returncode != 0:
            return []
        return [int(pid.strip()) for pid in result.stdout.splitlines() if pid.strip().isdigit()]


    # 先发 SIGTERM 后发 SIGKILL，确保 trinity_test 用户的所有进程退出
    def _terminate_user_processes(self):
        pids = self._get_user_pids()
        for pid in pids:
            try:
                os.kill(pid, signal.SIGTERM)
            except ProcessLookupError:
                continue
            except PermissionError:
                continue

        for _ in range(5):
            time.sleep(1)
            pids = self._get_user_pids()
            if not pids:
                return

        for pid in pids:
            try:
                os.kill(pid, signal.SIGKILL)
            except ProcessLookupError:
                continue
            except PermissionError:
                continue

        time.sleep(1)


    # 解析 trinity 日志，提取执行统计、crash 栈、可疑 syscall 等信息
    def _parse_trinity_log(self, content: str):
        lines = content.splitlines()
        summary = {
            '版本': '',
            '编译器': self.compiler,
            '测试结论': '未知',
            '随机种子': '',
            '启用syscall数': '',
            '禁用syscall数': '',
            '目标操作数': '',
            '实际执行syscall数': '',
            '成功次数': '',
            '失败次数': '',
            '失败率': '',
            '关键结论': '',
            '日志文件': self.log_path.name,
            'crash栈数量': 0,
            '可疑syscall线索数': 0,
        }

        version_match = re.search(r'^(Trinity[^\n]*)', content, re.MULTILINE)
        if version_match:
            summary['版本'] = version_match.group(1).strip()

        seed_match = re.search(r'Initial random seed:\s*(\d+)', content)
        if seed_match:
            summary['随机种子'] = seed_match.group(1)

        enabled_match = re.search(r'Enabled\s+(\d+)\s+syscalls\.\s+Disabled\s+(\d+)\s+syscalls\.', content)
        if enabled_match:
            summary['启用syscall数'] = int(enabled_match.group(1))
            summary['禁用syscall数'] = int(enabled_match.group(2))

        limit_match = re.search(r'Reached limit\s+(\d+)', content)
        if limit_match:
            summary['目标操作数'] = int(limit_match.group(1))

        ran_match = re.search(r'Ran\s+(\d+)\s+syscalls\.\s+Successes:\s+(\d+)\s+Failures:\s+(\d+)', content)
        if ran_match:
            total = int(ran_match.group(1))
            success = int(ran_match.group(2))
            failure = int(ran_match.group(3))
            summary['实际执行syscall数'] = total
            summary['成功次数'] = success
            summary['失败次数'] = failure
            summary['失败率'] = f'{(failure / total) * 100:.2f}%' if total else ''

        main_events = []
        for line in lines:
            if '[main]' not in line:
                continue
            if any(keyword in line for keyword in (
                "Couldn't",
                'Reached limit',
                'exit_reason',
                'Bailing',
                'Ran ',
                'Disabled',
                'Enabled',
            )):
                main_events.append(line.strip())

        if any('Reached limit' in line for line in main_events) and summary['实际执行syscall数']:
            summary['测试结论'] = '成功'
            summary['关键结论'] = '达到设定操作数上限，trinity正常完成'
        elif summary['实际执行syscall数']:
            summary['测试结论'] = '部分完成'
            summary['关键结论'] = '有执行结果，但未发现达到目标操作数的明确标记'
        else:
            summary['测试结论'] = '失败'
            summary['关键结论'] = '日志中未解析到最终执行统计'

        syscall_counter = Counter()
        errno_counter = Counter()
        child_counter = Counter()
        suspicious_syscalls = []
        child_syscall_history = defaultdict(lambda: deque(maxlen=25))
        crash_stacks = []
        repro_clues = []

        syscall_pattern = re.compile(r'^\[(child\d+):\d+\]\s+\[\d+\]\s+([a-zA-Z0-9_]+)\(')
        errno_pattern = re.compile(r'= -1 \(([^)]+)\)')
        child_line_pattern = re.compile(r'^\[(child\d+):(\d+)\]\s+\[(\d+)\]\s+')
        crash_start_keywords = (
            'BUG:',
            'Oops:',
            'Call Trace:',
            'Kernel panic',
            'general protection fault',
            'Unable to handle kernel',
            'Internal error:',
            'watchdog: BUG:',
            'RIP:',
            'pc :',
            'epc :',
            'lr :',
            'die at',
            'segfault at',
        )
        suspicious_errno_keywords = (
            'Permission denied',
            'Operation not permitted',
            'Transport endpoint is not connected',
            'Required key not available',
            'Bad address',
            'No such file or directory',
            'Connection reset by peer',
            'Broken pipe',
            'Input/output error',
            'Resource temporarily unavailable',
        )

        i = 0
        while i < len(lines):
            line = lines[i]
            syscall_match = syscall_pattern.match(line)
            if syscall_match:
                child_name, syscall_name = syscall_match.groups()
                child_counter[child_name] += 1
                syscall_counter[syscall_name] += 1
                child_syscall_history[child_name].append(line.strip())

            errno_match = errno_pattern.search(line)
            if errno_match:
                errno_name = errno_match.group(1)
                errno_counter[errno_name] += 1
                if any(keyword in errno_name for keyword in suspicious_errno_keywords):
                    child_info = child_line_pattern.match(line)
                    child_name = child_info.group(1) if child_info else ''
                    pid = child_info.group(2) if child_info else ''
                    syscall_name = ''
                    if syscall_match:
                        syscall_name = syscall_match.group(2)
                    elif child_name and child_syscall_history[child_name]:
                        last_line = child_syscall_history[child_name][-1]
                        last_match = syscall_pattern.match(last_line)
                        if last_match:
                            syscall_name = last_match.group(2)
                    suspicious_syscalls.append({
                        'child': child_name,
                        'pid': pid,
                        'syscall': syscall_name,
                        'errno': errno_name,
                        'line': line.strip(),
                        'context': '\n'.join(child_syscall_history[child_name]) if child_name else line.strip(),
                    })

            if any(keyword in line for keyword in crash_start_keywords):
                start = max(0, i - 8)
                end = min(len(lines), i + 30)
                block = []
                child_hint = ''
                for j in range(start, end):
                    current_line = lines[j]
                    if j > i and not current_line.strip():
                        break
                    block.append(current_line)
                    child_match = child_line_pattern.match(current_line)
                    if child_match and not child_hint:
                        child_hint = child_match.group(1)
                related_syscalls = list(child_syscall_history[child_hint]) if child_hint else []
                crash_stacks.append({
                    'start_line': i + 1,
                    'headline': line.strip(),
                    'child': child_hint,
                    'stack': '\n'.join(block).strip(),
                    'syscall_context': '\n'.join(related_syscalls).strip(),
                })
                i = end
                continue

            i += 1

        for child_name, history in child_syscall_history.items():
            if history:
                repro_clues.append({
                    'child': child_name,
                    'last_syscall_line': history[-1],
                    'recent_syscalls': '\n'.join(history),
                })

        repro_clues = sorted(
            repro_clues,
            key=lambda item: int(item['child'].replace('child', ''))
        )

        summary['crash栈数量'] = len(crash_stacks)
        summary['可疑syscall线索数'] = len(suspicious_syscalls)

        if crash_stacks:
            summary['测试结论'] = '疑似内核异常'
            summary['关键结论'] = '日志中发现疑似crash栈，请优先查看crash stacks和repro clues'
        elif any('Reached limit' in line for line in main_events) and summary['实际执行syscall数']:
            summary['测试结论'] = '成功'
            summary['关键结论'] = '达到设定操作数上限，未在trinity日志中发现明显crash栈'
        elif summary['实际执行syscall数']:
            summary['测试结论'] = '部分完成'
            summary['关键结论'] = '有执行结果，但未发现达到目标操作数的明确标记'
        else:
            summary['测试结论'] = '失败'
            summary['关键结论'] = '日志中未解析到最终执行统计'

        return summary, syscall_counter, errno_counter, child_counter, main_events, suspicious_syscalls, crash_stacks, repro_clues


    # 将解析结果写入包含 summary、syscall、crash 等多个 sheet 的 Excel 报告
    def _write_report(self):
        wb = Workbook()
        ws = wb.active
        ws.title = 'summary'

        summary, syscall_counter, errno_counter, child_counter, main_events, suspicious_syscalls, crash_stacks, repro_clues = self._parse_trinity_log(self.test_result)

        ws.append(['项目', '值'])
        for key, value in summary.items():
            ws.append([key, value])

        wb.create_sheet(title='main events')
        ws = wb['main events']
        ws.append(['序号', '关键事件'])
        if main_events:
            for index, line in enumerate(main_events, start=1):
                ws.append([index, line])
        else:
            ws.append([1, '未提取到关键主流程事件'])

        wb.create_sheet(title='top syscalls')
        ws = wb['top syscalls']
        ws.append(['排名', 'syscall', '出现次数'])
        if syscall_counter:
            for index, (name, count) in enumerate(syscall_counter.most_common(50), start=1):
                ws.append([index, name, count])
        else:
            ws.append([1, '未解析到syscall', 0])

        wb.create_sheet(title='error summary')
        ws = wb['error summary']
        ws.append(['排名', '错误类型', '出现次数'])
        if errno_counter:
            for index, (name, count) in enumerate(errno_counter.most_common(30), start=1):
                ws.append([index, name, count])
        else:
            ws.append([1, '未解析到错误返回', 0])

        wb.create_sheet(title='child activity')
        ws = wb['child activity']
        ws.append(['排名', '子进程', '记录条数'])
        if child_counter:
            for index, (name, count) in enumerate(child_counter.most_common(), start=1):
                ws.append([index, name, count])
        else:
            ws.append([1, '未解析到子进程活动', 0])

        wb.create_sheet(title='repro clues')
        ws = wb['repro clues']
        ws.append(['序号', 'child', '最后一条syscall', '最近syscall上下文'])
        if repro_clues:
            for index, item in enumerate(repro_clues[:80], start=1):
                ws.append([
                    index,
                    item['child'],
                    item['last_syscall_line'],
                    item['recent_syscalls'],
                ])
        else:
            ws.append([1, '', '未提取到可复现线索', ''])

        wb.create_sheet(title='suspicious syscalls')
        ws = wb['suspicious syscalls']
        ws.append(['序号', 'child', 'pid', 'syscall', 'errno', '原始日志', '上下文'])
        if suspicious_syscalls:
            for index, item in enumerate(suspicious_syscalls[:200], start=1):
                ws.append([
                    index,
                    item['child'],
                    item['pid'],
                    item['syscall'],
                    item['errno'],
                    item['line'],
                    item['context'],
                ])
        else:
            ws.append([1, '', '', '', '', '未提取到可疑syscall', ''])

        wb.create_sheet(title='crash stacks')
        ws = wb['crash stacks']
        ws.append(['序号', '起始行', 'child', 'crash摘要', '关联syscall上下文', 'crash栈'])
        if crash_stacks:
            for index, item in enumerate(crash_stacks, start=1):
                ws.append([
                    index,
                    item['start_line'],
                    item['child'],
                    item['headline'],
                    item['syscall_context'],
                    item['stack'],
                ])
        else:
            ws.append([1, '', '', '未发现明显crash栈', '', ''])

        wb.save(self.report_path)


    def pre_test(self):
        if not self.directory.exists():
            self.directory.mkdir(exist_ok=True,parents=True)
        user_exist = subprocess.run(
            "id trinity_test",
            shell=True,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL
        )
        # trinity_test用户若不存在则创建一个
        if user_exist.returncode != 0:
            try:
                subprocess.run(
                    "useradd -m trinity_test",
                    shell=True,
                    check=True,
                    stdout=subprocess.DEVNULL,
                    stderr=subprocess.PIPE
                )
            except subprocess.CalledProcessError as e:
                raise DefaultError(f"trinity测试出错:无法创建临时测试用户trinity_test.报错信息:{e.stderr.decode('utf-8')}")

        else:
            self._terminate_user_processes()
            del_add = subprocess.run(
                f"userdel trinity_test -r && useradd -m trinity_test",
                shell=True,
                stdout=subprocess.DEVNULL,
                stderr=subprocess.PIPE
            )
            if del_add.returncode != 0:
                print(f"trinity测试出错:无法创建临时测试用户trinity_test.报错信息:{del_add.stderr.decode('utf-8')}")
                sys.exit(1)

        if not self.user_home.exists():
            raise DefaultError(f"trinity测试出错:测试用户家目录不存在:{self.user_home}")

        try:
            subprocess.run(
                f"git clone https://gitee.com/April_Zhao/trinity_{self.compiler}.git",
                cwd=self.user_home,
                shell=True,check=True,
                stdout=subprocess.DEVNULL,
                stderr=subprocess.PIPE
            )
        except subprocess.CalledProcessError as e:
            raise GitCloneError(e.returncode,f'https://gitee.com/April_Zhao/trinity_{self.compiler}.git',e.stderr.decode('utf-8'))

        try:
            subprocess.run(
                f"./configure && make && make install",
                cwd=self.trinity_path,
                shell=True,check=True,
                stdout=subprocess.DEVNULL,
                stderr=subprocess.PIPE
            )
        except subprocess.CalledProcessError as e:
            raise DefaultError(f"trinity测试出错:configure和make失败.报错信息:{e.stderr.decode('utf-8')}")

        try:
            subprocess.run(
                f"chown -R trinity_test:trinity_test {self.trinity_path}",
                shell=True,
                check=True,
                stdout=subprocess.DEVNULL,
                stderr=subprocess.PIPE
            )
        except subprocess.CalledProcessError as e:
            raise DefaultError(f"trinity测试出错:trinity_{self.compiler}目录的权限设置失败.报错信息:{e.stderr.decode('utf-8')}")


    def run_test(self):
        try:
            trinity = subprocess.run(
                f"""su - trinity_test -c 'cd {self.trinity_path} && ./trinity -N 10000'""",
                shell=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE
            )
        except subprocess.CalledProcessError as e:
            raise DefaultError(f"trinity测试出错:configure和make失败.报错信息:{e.stderr.decode('utf-8')}")
        else:
            stdout = trinity.stdout.decode('utf-8', errors='replace')
            stderr = trinity.stderr.decode('utf-8', errors='replace')
            self.test_result = stdout
            if stderr.strip():
                self.test_result = f"{stdout}\n\n[stderr]\n{stderr}"
            with open(self.log_path, 'w', encoding='utf-8') as file:
                file.write(self.test_result)
            self._write_report()


    def post_test(self):
        self._terminate_user_processes()
        userdel = subprocess.run(
            "userdel trinity_test -r", # -r选项会删除用户的家目录
                shell=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE
        )
        if userdel.returncode != 0:
            print(
                "删除trinity的测试用户trinity_test失败,请手动执行[userdel trinity_test -r]."
                f"报错信息:{userdel.stderr.decode('utf-8')}"
            )


    def run(self):
        print("开始进行trinity测试")
        self.pre_test()
        try:
            self.run_test()
        finally:
            self.post_test()
        print("trinity测试结束")

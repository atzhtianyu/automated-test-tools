from openpyxl.workbook import Workbook
from openpyxl.styles import Alignment
from pystemd.systemd1 import Unit
from pathlib import Path
import re,os,shlex,statistics
import pymysql,time
import sys,subprocess,shutil

from .errors import DefaultError,RunError,SummaryError


class sysBench:
    def __init__(self, **kwargs):
        self.rpms = {'sysbench','mysql-server'}
        self.directory: Path = kwargs.get('saved_directory') / 'sysbench'
        self.test_result:str = ''
        self.prepare_succeeded = False
        self.table_size = int(kwargs.get('sysbench_table_size', 100000))
        self.tables = int(kwargs.get('sysbench_tables', 16))
        self.duration = int(kwargs.get('sysbench_time', 60))
        self.run_threads = int(kwargs.get('sysbench_threads', min(os.cpu_count() or 1, 4)))
        self.prepare_threads = int(kwargs.get('sysbench_prepare_threads', self.run_threads))
        self.command_prefix = (
            "sysbench --db-driver=mysql --mysql-host=127.0.0.1 "
            "--mysql-port=3306 --mysql-user=root --mysql-password=123456 "
            f"--mysql-db=sysbench --table_size={self.table_size} --tables={self.tables} --time={self.duration} "
        )
        self.prepare_command = f"{self.command_prefix}--threads={self.prepare_threads} --report-interval=1 oltp_read_write prepare"
        self.run_command = f"{self.command_prefix}--threads={self.run_threads} --report-interval=1 oltp_read_write run"
        self.cleanup_command = f"{self.command_prefix}--threads={self.prepare_threads} --report-interval=1 oltp_read_write cleanup"


    # 将 part/total 格式化为百分比字符串，分母为零时返回 '/'
    def _format_percentage(self, part: int | float, total: int | float):
        if not total:
            return '/'
        return f"{part / total:.4%}"


    # 执行命令并同时捕获 stdout 和 stderr，返回 CompletedProcess
    def _run_capture_command(self, command: list[str]):
        return subprocess.run(
            command,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
        )


    # 将命令、返回码及 stdout/stderr 以结构化格式写入日志文件
    def _write_command_log(self, log_path: Path, command: str, returncode: int, stdout: str = '', stderr: str = ''):
        sections = [
            f"[command]\n{command}",
            f"[returncode]\n{returncode}",
            f"[stdout]\n{stdout.strip() if stdout else '<empty>'}",
            f"[stderr]\n{stderr.strip() if stderr else '<empty>'}",
        ]
        log_path.write_text('\n\n'.join(sections) + '\n', encoding='utf-8')


    # 轮询等待 systemd 服务进入 active 状态，超时返回 False
    def _wait_for_service_active(self, service_name: str, timeout: int = 30):
        deadline = time.time() + timeout
        last_state = ''
        while time.time() < deadline:
            result = self._run_capture_command(['systemctl', 'is-active', service_name])
            last_state = (result.stdout or result.stderr).strip()
            if result.returncode == 0 and last_state == 'active':
                return True, last_state
            time.sleep(1)
        return False, last_state or 'unknown'


    # 收集 mysqld 的服务状态、journalctl 和错误日志，用于启动失败诊断
    def _collect_mysqld_diagnostics(self):
        sections = []

        try:
            self.mysqld.load(force=True)
            active_state = self.mysqld.Unit.ActiveState.decode('utf-8', errors='ignore')
            sub_state = self.mysqld.Unit.SubState.decode('utf-8', errors='ignore')
            sections.append(
                "[pystemd]\n"
                f"ActiveState: {active_state}\n"
                f"SubState: {sub_state}\n"
            )
        except Exception as e:
            sections.append(f"[pystemd]\n读取服务状态失败: {e}\n")

        diagnostics_commands = [
            ['systemctl', 'status', 'mysqld.service', '--no-pager', '-l'],
            ['journalctl', '-u', 'mysqld.service', '-n', '80', '--no-pager'],
            ['tail', '-n', '80', '/var/log/mysql/mysqld.log'],
        ]
        for command in diagnostics_commands:
            result = self._run_capture_command(command)
            output = (result.stdout + result.stderr).strip() or '<empty>'
            sections.append(
                f"[{' '.join(command)}]\n"
                f"returncode: {result.returncode}\n"
                f"{output}\n"
            )

        return '\n'.join(sections)


    # 启动 mysqld.service，失败时保存诊断日志并抛出异常
    def _start_mysqld(self):
        self.mysqld:Unit = Unit('mysqld.service',_autoload=True)
        start_result = self._run_capture_command(['systemctl', 'start', 'mysqld.service'])
        active, last_state = self._wait_for_service_active('mysqld.service')
        if active:
            self.mysqld.load(force=True)
            return

        start_log = self.directory / 'mysqld_start.log'
        details = [
            "[systemctl start mysqld.service]",
            f"returncode: {start_result.returncode}",
            start_result.stdout.strip(),
            start_result.stderr.strip(),
            "",
            f"[poll result]\nlast_state: {last_state}",
            "",
            self._collect_mysqld_diagnostics(),
        ]
        start_log.write_text('\n'.join(part for part in details if part is not None), encoding='utf-8')
        raise DefaultError(
            f"sysbench测试出错.开启mysqld.service失败,最后状态:{last_state}.详细信息请查看:{start_log}"
        )


    # 从 sysbench 输出中解析逐秒趋势数据（TPS、QPS、延迟等）
    def _parse_timeline_rows(self):
        pattern = re.compile(
            r"^\[\s*(?P<time_sec>\d+)s \] thds: (?P<threads>\d+) tps: (?P<tps>[\d.]+) "
            r"qps: (?P<qps>[\d.]+) \(r/w/o: (?P<read_qps>[\d.]+)/(?P<write_qps>[\d.]+)/(?P<other_qps>[\d.]+)\) "
            r"lat \(ms,95%\): (?P<latency_95_ms>[\d.]+) err/s: (?P<err_per_sec>[\d.]+) reconn/s: (?P<reconn_per_sec>[\d.]+)$",
            re.MULTILINE,
        )

        rows = []
        for match in pattern.finditer(self.test_result):
            rows.append({
                'time_sec': int(match.group('time_sec')),
                'threads': int(match.group('threads')),
                'tps': float(match.group('tps')),
                'qps': float(match.group('qps')),
                'read_qps': float(match.group('read_qps')),
                'write_qps': float(match.group('write_qps')),
                'other_qps': float(match.group('other_qps')),
                'latency_95_ms': float(match.group('latency_95_ms')),
                'err_per_sec': float(match.group('err_per_sec')),
                'reconn_per_sec': float(match.group('reconn_per_sec')),
            })

        if not rows:
            raise ValueError('sysbench日志中未找到逐秒趋势数据')
        return rows


    # 将 sysbench 命令行字符串解析为参数列表和测试脚本、动作字段
    def _parse_command_parameters(self, command: str):
        tokens = shlex.split(command)
        parameters = []
        test_script = ''
        action = ''

        index = 1
        while index < len(tokens):
            token = tokens[index]
            if token.startswith('--'):
                option = token[2:]
                value = ''
                if '=' in option:
                    option, value = option.split('=', 1)
                elif index + 1 < len(tokens) and not tokens[index + 1].startswith('--'):
                    value = tokens[index + 1]
                    index += 1
                parameters.append({'option': option, 'value': value})
            elif not test_script:
                test_script = token
            elif not action:
                action = token
            index += 1

        return parameters, test_script, action


    # 执行 sysbench 命令并保存日志
    def _run_sysbench_command(self, command: str, log_filename: str):
        result = subprocess.run(
            command, shell=True,
            stdout=subprocess.PIPE, stderr=subprocess.PIPE,
        )
        self._write_command_log(
            self.directory / log_filename,
            command, result.returncode,
            result.stdout.decode('utf-8', errors='replace'),
            result.stderr.decode('utf-8', errors='replace'),
        )
        return result


    def pre_test(self):
        if self.directory.exists():
            shutil.rmtree(self.directory)
        self.directory.mkdir(parents=True)
        self._start_mysqld()

        try:
            self.conn = pymysql.connect(
                host='localhost',
                port=3306,
                user='root',
                passwd='',
            )
        except Exception as e:
            self.conn = pymysql.connect(
                host='localhost',
                port=3306,
                user='root',
                passwd='123456',
            )
        cursor = self.conn.cursor()
        cursor.execute("ALTER USER 'root'@'localhost' IDENTIFIED BY '123456';")
        cursor.execute("DROP DATABASE IF EXISTS sysbench;")
        cursor.execute("CREATE DATABASE IF NOT EXISTS sysbench;")
        self.conn.commit()
        cursor.close()
        self.conn.close()

        # 清理测试数据
        self._run_sysbench_command(self.cleanup_command, 'sysbench_cleanup.log')

        # 准备测试数据和表
        prepare = self._run_sysbench_command(self.prepare_command, 'sysbench_prepare.log')
        if prepare.returncode != 0:
            raise DefaultError(
                f"sysbench测试出错.准备测试数据和表失败,详细信息请查看:{self.directory / 'sysbench_prepare.log'}"
            )
        self.prepare_succeeded = True


    def run_test(self):
        sysbench_run = self._run_sysbench_command(self.run_command, 'sysbench.log')
        stdout = sysbench_run.stdout.decode('utf-8', errors='replace')
        stderr = sysbench_run.stderr.decode('utf-8', errors='replace')
        self.test_result = stdout

        if sysbench_run.returncode != 0:
            error_message = stderr.strip() or f"详细信息请查看:{self.directory / 'sysbench.log'}"
            if stderr.strip():
                error_message = f"{error_message}\n详细信息请查看:{self.directory / 'sysbench.log'}"
            raise RunError(sysbench_run.returncode, error_message)


    def result2summary(self):
        timeline_rows = self._parse_timeline_rows()
        command_parameters, test_script, command_action = self._parse_command_parameters(self.run_command)
        parameter_map = {item['option']: item['value'] for item in command_parameters}

        wb = Workbook()
        ws_summary = wb.active
        ws_summary.title = 'summary'
        ws_summary.append(['metric', 'value', 'note'])

        thread_count = re.search(r"Number of threads:\s*(\d+)", self.test_result).group(1)
        report_interval = re.search(r"Report intermediate results every\s*(\d+)\s*second", self.test_result).group(1)
        read_select = int(re.search(r"read:\s*(\d+)",self.test_result).group(1))
        write_select = int(re.search(r"write:\s*(\d+)",self.test_result).group(1))
        other_select = int(re.search(r"other:\s*(\d+)",self.test_result).group(1))
        total_select = int(re.search(r"total:\s*(\d+)",self.test_result).group(1))
        transactions = re.search(r"transactions:\s*(\d+)\s*\((\d+\.\d+) per sec\.\)",self.test_result).groups()
        query_count = re.search(r"queries:\s*(\d+)\s*\((\d+\.\d+) per sec\.\)",self.test_result).groups()
        ignore_errors = re.search(r"ignored errors:\s*(\d+)\s*\((\d+\.\d+) per sec\.\)",self.test_result).groups()
        reconnects = re.search(r"reconnects:\s*(\d+)\s*\((\d+\.\d+) per sec\.\)",self.test_result).groups()
        total_time = re.search(r"total time:\s*(\d+\.\d+)s",self.test_result).group(1)
        total_number_of_events = re.search(r"total number of events:\s*(\d+)",self.test_result).group(1)
        min_latency = re.search(r"min:\s*(\d+\.\d+)",self.test_result).group(1)
        avg_latency = re.search(r"avg:\s*(\d+\.\d+)",self.test_result).group(1)
        max_latency = re.search(r"max:\s*(\d+\.\d+)",self.test_result).group(1)
        percentile_95th = re.search(r"95th percentile:\s*(\d+\.\d+)",self.test_result).group(1)
        latency_sum = re.search(r"sum:\s*(\d+\.\d+)",self.test_result).group(1)
        events_avg,events_stddev = re.search(r"events \(avg/stddev\):\s*(\d+\.\d+)/(\d+\.\d+)",self.test_result).groups()
        execution_time_avg,execution_time_stddev = re.search(r"execution time \(avg/stddev\):\s*(\d+\.\d+)/(\d+\.\d+)",self.test_result).groups()

        average_tps = round(sum(row['tps'] for row in timeline_rows) / len(timeline_rows), 2)
        median_tps = round(statistics.median(row['tps'] for row in timeline_rows), 2)
        peak_tps = max(timeline_rows, key=lambda row: row['tps'])
        lowest_tps = min(timeline_rows, key=lambda row: row['tps'])
        average_qps = round(sum(row['qps'] for row in timeline_rows) / len(timeline_rows), 2)
        median_qps = round(statistics.median(row['qps'] for row in timeline_rows), 2)
        peak_qps = max(timeline_rows, key=lambda row: row['qps'])
        average_latency95 = round(sum(row['latency_95_ms'] for row in timeline_rows) / len(timeline_rows), 2)
        peak_latency = max(timeline_rows, key=lambda row: row['latency_95_ms'])
        error_intervals = sum(1 for row in timeline_rows if row['err_per_sec'] > 0)
        reconnect_intervals = sum(1 for row in timeline_rows if row['reconn_per_sec'] > 0)

        summary_rows = [
            ['test command', self.run_command, 'sysbench实际执行命令'],
            ['test script', test_script, 'sysbench内置Lua脚本'],
            ['command action', command_action, '当前命令阶段'],
            ['db driver', parameter_map.get('db-driver', ''), '数据库驱动'],
            ['database', parameter_map.get('mysql-db', ''), '目标数据库名'],
            ['table size', int(parameter_map['table_size']) if parameter_map.get('table_size', '').isdigit() else parameter_map.get('table_size', ''), '单表记录数'],
            ['tables', int(parameter_map['tables']) if parameter_map.get('tables', '').isdigit() else parameter_map.get('tables', ''), '测试表数量'],
            ['configured duration(s)', int(parameter_map['time']) if parameter_map.get('time', '').isdigit() else parameter_map.get('time', ''), 'sysbench配置的run时长'],
            ['configured threads', int(thread_count), 'sysbench报告的线程数'],
            ['report interval(s)', int(report_interval), '逐秒采样间隔'],
            ['timeline samples', len(timeline_rows), '逐秒趋势点数量'],
            ['average interval tps', average_tps, '逐秒TPS均值'],
            ['median interval tps', median_tps, '逐秒TPS中位数'],
            ['peak interval tps', peak_tps['tps'], f"{peak_tps['time_sec']}s"],
            ['lowest interval tps', lowest_tps['tps'], f"{lowest_tps['time_sec']}s"],
            ['average interval qps', average_qps, '逐秒QPS均值'],
            ['median interval qps', median_qps, '逐秒QPS中位数'],
            ['peak interval qps', peak_qps['qps'], f"{peak_qps['time_sec']}s"],
            ['average interval latency95(ms)', average_latency95, '逐秒95分位延迟均值'],
            ['peak interval latency95(ms)', peak_latency['latency_95_ms'], f"{peak_latency['time_sec']}s"],
            ['timeline error intervals', error_intervals, 'err/s大于0的采样点数量'],
            ['timeline reconnect intervals', reconnect_intervals, 'reconn/s大于0的采样点数量'],
            ['total time(s)', float(total_time), '最终汇总总耗时'],
            ['transactions', int(transactions[0]), '最终汇总事务总数'],
            ['transactions per sec', float(transactions[1]), '最终汇总TPS'],
            ['queries', int(query_count[0]), '最终汇总查询总数'],
            ['queries per sec', float(query_count[1]), '最终汇总QPS'],
            ['ignored errors', int(ignore_errors[0]), '最终汇总忽略错误数'],
            ['reconnects', int(reconnects[0]), '最终汇总重连次数'],
            ['avg latency(ms)', float(avg_latency), '最终汇总平均延迟'],
            ['95th latency(ms)', float(percentile_95th), '最终汇总95分位延迟'],
        ]
        for row in summary_rows:
            ws_summary.append(row)

        ws_parameters = wb.create_sheet(title='parameters')
        ws_parameters.append(['option', 'value'])
        ws_parameters.append(['test script', test_script])
        ws_parameters.append(['action', command_action])
        for item in command_parameters:
            ws_parameters.append([item['option'], item['value']])

        ws_timeline = wb.create_sheet(title='timeline')
        ws_timeline.append([
            'time(s)', 'threads', 'tps', 'qps',
            'read qps', 'write qps', 'other qps',
            'latency95(ms)', 'err/s', 'reconn/s',
        ])
        for row in timeline_rows:
            ws_timeline.append([
                row['time_sec'],
                row['threads'],
                row['tps'],
                row['qps'],
                row['read_qps'],
                row['write_qps'],
                row['other_qps'],
                row['latency_95_ms'],
                row['err_per_sec'],
                row['reconn_per_sec'],
            ])

        ws = wb.create_sheet(title='final_stats')
        ws.append(['category', 'metric', 'value', 'note'])

        final_rows = [
            ['sql statistics', 'read queries', read_select, self._format_percentage(read_select, total_select)],
            ['sql statistics', 'write queries', write_select, self._format_percentage(write_select, total_select)],
            ['sql statistics', 'other queries', other_select, self._format_percentage(other_select, total_select)],
            ['sql statistics', 'total queries', total_select, '/'],
            ['sql statistics', 'transactions', int(transactions[0]), '/'],
            ['sql statistics', 'transactions per sec', float(transactions[1]), '/'],
            ['sql statistics', 'queries per sec', float(query_count[1]), '/'],
            ['sql statistics', 'ignored errors per sec', float(ignore_errors[1]), '/'],
            ['sql statistics', 'reconnects per sec', float(reconnects[1]), '/'],
            ['general statistics', 'total time(s)', float(total_time), '/'],
            ['general statistics', 'total number of events', int(total_number_of_events), '/'],
            ['latency', 'min(ms)', float(min_latency), '/'],
            ['latency', 'avg(ms)', float(avg_latency), '/'],
            ['latency', 'max(ms)', float(max_latency), '/'],
            ['latency', '95th percentile(ms)', float(percentile_95th), '/'],
            ['latency', 'sum(ms)', float(latency_sum), '/'],
            ['threads fairness', 'events avg', float(events_avg), '/'],
            ['threads fairness', 'events stddev', float(events_stddev), '越小越好'],
            ['threads fairness', 'execution time avg(s)', float(execution_time_avg), '/'],
            ['threads fairness', 'execution time stddev', float(execution_time_stddev), '越小越好'],
        ]
        for row in final_rows:
            ws.append(row)

        for sheet in (ws_summary, ws_parameters, ws_timeline, ws):
            sheet.freeze_panes = 'A2'
            for row in sheet.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(vertical='top', wrap_text=True)

        ws_summary.column_dimensions['A'].width = 24
        ws_summary.column_dimensions['B'].width = 22
        ws_summary.column_dimensions['C'].width = 24

        ws_parameters.column_dimensions['A'].width = 24
        ws_parameters.column_dimensions['B'].width = 26

        for column in ('A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J'):
            ws_timeline.column_dimensions[column].width = 16

        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 24
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18

        wb.save(self.directory / 'sysbench.xlsx')



    def post_test(self):
        if self.prepare_succeeded:
            sysbench_clean = subprocess.run(
                self.cleanup_command,
                shell=True,
                stdout=subprocess.DEVNULL,
                stderr=subprocess.PIPE,
            )
            if sysbench_clean.returncode != 0:
                print(f"sysbench测试.清理测试数据失败,报错信息:{sysbench_clean.stderr.decode('utf-8')}")

        if hasattr(self, 'mysqld'):
            try:
                self.mysqld.Unit.Stop(b'replace')
            except Exception as e:
                print(f"sysbench测试.停止mysqld.service失败,报错信息:{str(e)}")


    def run(self):
        print('开始进行sysbench测试')
        need_cleanup = False
        try:
            self.pre_test()
            need_cleanup = True
            self.run_test()
            try:
                self.result2summary()
            except Exception as e:
                logFile = self.directory / 'sysbench_summary_error.log'
                with open(logFile, 'w') as log:
                    log.write(str(e))
                raise SummaryError(logFile)
        finally:
            if need_cleanup or hasattr(self, 'mysqld'):
                self.post_test()
        print('sysbench测试结束')
